from datetime import datetime
import os
import re
import csv
import io
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.staticfiles import finders
from django.views.decorators.http import require_http_methods
from django.db.models import Q, Count, Avg, Min, Max, Case, When, FloatField, F
from django.db.models.functions import Cast
from django.urls import reverse
from django.db import IntegrityError
from django.core.cache import cache
from django.views.decorators.cache import cache_page
from django.http import HttpResponse, FileResponse, Http404
from django.conf import settings
from rest_framework import viewsets, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import (
    School, User, ClassSection, Subject, GradingScale, StudentEnrollment,
    GradingPeriod, Grade, Attendance, UserApplication, SchoolProfile,
    SupportTicket, ReportCard, ReportTemplate
)
from .serializers import (
    SchoolSerializer, UserSerializer, ClassSectionSerializer,
    SubjectSerializer, GradingScaleSerializer, StudentEnrollmentSerializer,
    GradingPeriodSerializer, GradeSerializer, AttendanceSerializer,
    SchoolProfileSerializer, SupportTicketSerializer, ReportCardSerializer
)
from .mixins import StandardViewSet, StudentOwnerFilterMixin, ExportMixin
from .utils import (
    PermissionHelper, AnalyticsHelper, ValidationHelper,
    ExcelExporter, PDFExporter, CSVExporter
)
from authentication.permissions import (
    IsSuperAdmin, IsSchoolAdmin, IsSchoolMember, IsOwnerOrSchoolAdmin,
    IsTeacher, IsStudent, IsStudentOwner, IsTeacherOrAdmin
)


# ViewSets using StandardViewSet base for code reuse
class SchoolViewSet(viewsets.ModelViewSet):
    queryset = School.objects.all()
    serializer_class = SchoolSerializer
    permission_classes = [IsSuperAdmin]


class UserViewSet(StandardViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsSchoolAdmin]


class ClassSectionViewSet(StandardViewSet):
    queryset = ClassSection.objects.all()
    serializer_class = ClassSectionSerializer
    permission_classes = [IsTeacherOrAdmin]


class SubjectViewSet(StandardViewSet):
    queryset = Subject.objects.all()
    serializer_class = SubjectSerializer
    permission_classes = [IsTeacherOrAdmin]


class GradingScaleViewSet(StandardViewSet):
    queryset = GradingScale.objects.all()
    serializer_class = GradingScaleSerializer
    permission_classes = [IsTeacherOrAdmin]


class GradingPeriodViewSet(StandardViewSet):
    queryset = GradingPeriod.objects.all()
    serializer_class = GradingPeriodSerializer
    permission_classes = [IsTeacherOrAdmin]


class StudentEnrollmentViewSet(StandardViewSet, StudentOwnerFilterMixin):
    queryset = StudentEnrollment.objects.all()
    serializer_class = StudentEnrollmentSerializer
    permission_classes = [IsTeacherOrAdmin]

    def get_permissions(self):
        if self.action in ['retrieve', 'list']:
            return [IsStudent(), IsStudentOwner()]
        return [IsTeacherOrAdmin()]


class GradeViewSet(StandardViewSet, StudentOwnerFilterMixin):
    queryset = Grade.objects.all()
    serializer_class = GradeSerializer
    permission_classes = [IsTeacherOrAdmin]

    def get_permissions(self):
        if self.action in ['retrieve', 'list']:
            return [IsStudent(), IsStudentOwner()]
        return [IsTeacherOrAdmin()]


class AttendanceViewSet(StandardViewSet, StudentOwnerFilterMixin):
    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer
    permission_classes = [IsTeacherOrAdmin]

    def get_permissions(self):
        if self.action in ['retrieve', 'list']:
            return [IsStudent(), IsStudentOwner()]
        return [IsTeacherOrAdmin()]


class SchoolProfileViewSet(StandardViewSet):
    queryset = SchoolProfile.objects.all()
    serializer_class = SchoolProfileSerializer
    permission_classes = [IsSchoolAdmin]


class SupportTicketViewSet(StandardViewSet):
    queryset = SupportTicket.objects.all()
    serializer_class = SupportTicketSerializer
    permission_classes = [IsSchoolMember]

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        
        if user.role == 'super_admin':
            return SupportTicket.objects.all()
        elif user.role == 'admin':
            return SupportTicket.objects.filter(school=user.school)
        else:
            return SupportTicket.objects.filter(created_by=user)


class ReportCardViewSet(StandardViewSet, StudentOwnerFilterMixin):
    queryset = ReportCard.objects.all()
    serializer_class = ReportCardSerializer
    permission_classes = [IsTeacherOrAdmin]

    def get_queryset(self):
        queryset = ReportCard.objects.all()
        user = self.request.user
        
        if user.role == 'super_admin':
            return queryset
        elif user.role == 'admin':
            return queryset.filter(school=user.school)
        elif user.role == 'teacher':
            student_ids = StudentEnrollment.objects.filter(
                class_section__teacher=user
            ).values_list('student_id', flat=True)
            return queryset.filter(student_id__in=student_ids, school=user.school)
        else:
            return queryset.filter(student=user)

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [IsStudent(), IsStudentOwner()]
        return [IsTeacherOrAdmin()]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def search_api_view(request):
    """API endpoint for global search functionality"""
    query = request.GET.get('q', '').strip()

    if not query or len(query) < 2:
        return Response({'results': []})

    # Sanitize query to prevent injection attacks
    import re
    query = re.sub(r'[^\w\s\-@.]', '', query)
    query = query.strip()[:100]  # Limit query length

    if not query:
        return Response({'results': []})

    results = []

    # Search users with proper parameterization
    users = User.objects.filter(
        Q(username__icontains=query) |
        Q(first_name__icontains=query) |
        Q(last_name__icontains=query) |
        Q(email__icontains=query)
    ).select_related('school')

    # Filter by school permissions
    if request.user.role == 'super_admin':
        pass  # Can see all users
    elif request.user.role == 'admin':
        users = users.filter(school=request.user.school)
    elif request.user.role == 'teacher':
        # Teachers can see students in their classes and other staff at their school
        teacher_students = StudentEnrollment.objects.filter(
            class_section__teacher=request.user
        ).values_list('student_id', flat=True)
        users = users.filter(
            Q(school=request.user.school) |
            Q(id__in=teacher_students)
        )
    else:
        # Students can only see their own information
        users = users.filter(id=request.user.id)

    users = users[:5]  # Limit results
    for user in users:
        results.append({
            'title': user.get_full_name() or user.username,
            'subtitle': f"{user.get_role_display()} • {user.school.name if user.school else 'No School'}",
            'url': f'/users/{user.id}/update/',
            'icon': 'person-fill' if user.role == 'student' else 'person-badge-fill',
            'type': 'user'
        })

    # Search class sections with proper parameterization
    classes = ClassSection.objects.filter(
        Q(name__icontains=query) |
        Q(grade_level__icontains=query)
    ).select_related('school')

    if request.user.role == 'super_admin':
        pass  # Can see all classes
    elif request.user.role in ['admin', 'teacher']:
        classes = classes.filter(school=request.user.school)
    else:
        # Students can only see their own classes
        student_classes = StudentEnrollment.objects.filter(
            student=request.user
        ).values_list('class_section_id', flat=True)
        classes = classes.filter(id__in=student_classes)

    classes = classes[:3]  # Limit results
    for cls in classes:
        results.append({
            'title': cls.name,
            'subtitle': f"Grade {cls.grade_level} • {cls.school.name}",
            'url': f'/class-sections/{cls.id}/update/',
            'icon': 'mortarboard-fill',
            'type': 'class'
        })

    # Search subjects with proper parameterization
    subjects = Subject.objects.filter(
        Q(name__icontains=query) |
        Q(code__icontains=query)
    ).select_related('school')

    if request.user.role == 'super_admin':
        pass  # Can see all subjects
    elif request.user.role in ['admin', 'teacher']:
        subjects = subjects.filter(school=request.user.school)
    else:
        # Students can only see subjects they're enrolled in
        student_subjects = Grade.objects.filter(
            student=request.user
        ).values_list('subject_id', flat=True).distinct()
        subjects = subjects.filter(id__in=student_subjects)

    subjects = subjects[:3]  # Limit results
    for subject in subjects:
        results.append({
            'title': subject.name,
            'subtitle': f"{subject.code} • {subject.school.name}",
            'url': f'/subjects/{subject.id}/update/',
            'icon': 'book-fill',
            'type': 'subject'
        })

    return Response({'results': results})



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def sync_view(request):
    last_sync_str = request.GET.get('last_sync')
    school_id = request.GET.get('school_id')  # Allow explicit school_id for offline sync

    if not last_sync_str:
        return Response({'error': 'last_sync parameter required'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        last_sync = datetime.fromisoformat(last_sync_str.replace('Z', '+00:00'))
    except ValueError:
        return Response({'error': 'Invalid last_sync format'}, status=status.HTTP_400_BAD_REQUEST)

    # Determine school context - use explicit school_id if provided, otherwise use request.school
    school_context = None
    if school_id:
        try:
            school_context = School.objects.get(id=school_id)
        except School.DoesNotExist:
            return Response({'error': 'Invalid school_id'}, status=status.HTTP_400_BAD_REQUEST)
    elif request.school:
        school_context = request.school

    data = {}
    models = [
        (School, SchoolSerializer),
        (User, UserSerializer),
        (ClassSection, ClassSectionSerializer),
        (Subject, SubjectSerializer),
        (GradingScale, GradingScaleSerializer),
        (GradingPeriod, GradingPeriodSerializer),
        (StudentEnrollment, StudentEnrollmentSerializer),
        (Grade, GradeSerializer),
        (Attendance, AttendanceSerializer),
    ]

    for model, serializer_class in models:
        queryset = model.objects.filter(updated_at__gt=last_sync)
        if hasattr(model, 'school') and school_context:
            queryset = queryset.filter(school=school_context)
        elif model == School and request.user.role != 'super_admin':
            queryset = queryset.none()
        elif model == User and request.user.role != 'super_admin':
            queryset = queryset.filter(school=school_context) if school_context else queryset.none()

        serializer = serializer_class(queryset, many=True)
        data[model.__name__.lower()] = serializer.data

    # Include user info and school context in response
    data['_meta'] = {
        'user_id': request.user.id,
        'school_id': school_context.id if school_context else None,
        'last_sync': datetime.now().isoformat(),
        'sync_timestamp': int(datetime.now().timestamp() * 1000)
    }

    return Response(data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def push_sync_view(request):
    """Accept batched changes from clients and apply them to the server.

    Payload example:
    {
        "grades": [{"action":"create","data":{...}}, {"action":"update","data":{...}}]
    }
    """
    payload = request.data or {}
    models = {
        'school': (School, SchoolSerializer),
        'user': (User, UserSerializer),
        'classsection': (ClassSection, ClassSectionSerializer),
        'subject': (Subject, SubjectSerializer),
        'gradingscale': (GradingScale, GradingScaleSerializer),
        'gradingperiod': (GradingPeriod, GradingPeriodSerializer),
        'studentenrollment': (StudentEnrollment, StudentEnrollmentSerializer),
        'grade': (Grade, GradeSerializer),
        'attendance': (Attendance, AttendanceSerializer),
    }

    result = {'created': {}, 'updated': {}, 'deleted': {}, 'errors': {}, 'conflicts': {}}

    for key, items in payload.items():
        model_key = key.rstrip('s').lower()
        if model_key not in models:
            continue
        Model, Serializer = models[model_key]
        result['created'].setdefault(key, [])
        result['updated'].setdefault(key, [])
        result['deleted'].setdefault(key, [])
        result['errors'].setdefault(key, [])
        result['conflicts'].setdefault(key, [])

        for entry in items or []:
            action = entry.get('action', 'update')
            data = entry.get('data') or {}
            try:
                if action == 'create':
                    ser = Serializer(data=data)
                    ser.is_valid(raise_exception=True)
                    obj = ser.save()
                    result['created'][key].append(Serializer(obj).data)

                elif action == 'update':
                    obj_id = data.get('id')
                    if not obj_id:
                        raise ValueError('Missing id for update')
                    try:
                        obj = Model.objects.get(id=obj_id)
                    except Model.DoesNotExist:
                        ser = Serializer(data=data)
                        ser.is_valid(raise_exception=True)
                        obj = ser.save()
                        result['created'][key].append(Serializer(obj).data)
                    else:
                        # Basic server-wins conflict detection
                        incoming_updated_at = None
                        try:
                            incoming_updated_at = data.get('updated_at')
                            if incoming_updated_at:
                                incoming_updated_at = datetime.fromisoformat(incoming_updated_at.replace('Z', '+00:00'))
                        except Exception:
                            incoming_updated_at = None

                        server_updated_at = getattr(obj, 'updated_at', None)

                        if server_updated_at and incoming_updated_at and server_updated_at > incoming_updated_at:
                            result['conflicts'][key].append({
                                'id': obj_id,
                                'reason': 'server_newer',
                                'server_updated_at': server_updated_at.isoformat(),
                                'incoming_updated_at': data.get('updated_at')
                            })
                        else:
                            ser = Serializer(obj, data=data, partial=True)
                            ser.is_valid(raise_exception=True)
                            obj = ser.save()
                            result['updated'][key].append(Serializer(obj).data)

                elif action == 'delete':
                    obj_id = data.get('id')
                    if not obj_id:
                        raise ValueError('Missing id for delete')
                    try:
                        obj = Model.objects.get(id=obj_id)
                        obj.delete()
                        result['deleted'][key].append({'id': obj_id})
                    except Model.DoesNotExist:
                        pass
                else:
                    result['errors'][key].append({'entry': entry, 'error': 'unknown action'})
            except Exception as e:
                result['errors'][key].append({'entry': entry, 'error': str(e)})

    return Response(result)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def offline_sync_batch_view(request):
    """
    Advanced offline sync endpoint for batched operations with conflict detection.
    Handles create, update, delete operations and returns detailed results.
    """
    payload = request.data or {}
    school_context = getattr(request, 'school', None)
    
    if not school_context and request.user.school:
        school_context = request.user.school

    result = {
        'success': True,
        'timestamp': datetime.now().isoformat(),
        'created': {},
        'updated': {},
        'deleted': {},
        'errors': {},
        'conflicts': {}
    }

    # Model mapping
    models = {
        'grades': (Grade, GradeSerializer),
        'attendance': (Attendance, AttendanceSerializer),
        'classsections': (ClassSection, ClassSectionSerializer),
        'subjects': (Subject, SubjectSerializer),
        'users': (User, UserSerializer),
        'reportcards': (ReportCard, ReportCardSerializer),
    }

    for model_key, items in payload.items():
        if model_key == '_meta' or not isinstance(items, list):
            continue

        model_key_lower = model_key.lower()
        if model_key_lower not in models:
            continue

        Model, Serializer = models[model_key_lower]
        result['created'].setdefault(model_key, [])
        result['updated'].setdefault(model_key, [])
        result['deleted'].setdefault(model_key, [])
        result['errors'].setdefault(model_key, [])
        result['conflicts'].setdefault(model_key, [])

        for entry in items:
            action = entry.get('action', 'update')
            data = entry.get('data') or {}
            object_id = data.get('id')

            try:
                # Enforce school context for security
                if hasattr(Model, 'school') and school_context:
                    data['school'] = school_context.id

                if action == 'create':
                    serializer = Serializer(data=data)
                    if serializer.is_valid():
                        obj = serializer.save()
                        result['created'][model_key].append(Serializer(obj).data)
                    else:
                        result['errors'][model_key].append({
                            'id': None,
                            'action': action,
                            'errors': serializer.errors
                        })

                elif action == 'update':
                    if not object_id:
                        result['errors'][model_key].append({
                            'id': None,
                            'action': action,
                            'errors': {'detail': 'Missing id for update'}
                        })
                        continue

                    try:
                        obj = Model.objects.get(id=object_id)
                        
                        # Check for conflict
                        incoming_updated_at = None
                        try:
                            if 'updated_at' in data:
                                incoming_updated_at = datetime.fromisoformat(
                                    data['updated_at'].replace('Z', '+00:00')
                                ) if isinstance(data['updated_at'], str) else data['updated_at']
                        except (ValueError, AttributeError, TypeError):
                            pass

                        server_updated_at = getattr(obj, 'updated_at', None)

                        # Detect conflict: server has newer version
                        if (server_updated_at and incoming_updated_at and 
                            server_updated_at > incoming_updated_at):
                            result['conflicts'][model_key].append({
                                'id': object_id,
                                'reason': 'server_newer',
                                'server_updated_at': server_updated_at.isoformat(),
                                'client_updated_at': data.get('updated_at'),
                                'server_data': Serializer(obj).data
                            })
                        else:
                            # No conflict, proceed with update
                            serializer = Serializer(obj, data=data, partial=True)
                            if serializer.is_valid():
                                obj = serializer.save()
                                result['updated'][model_key].append(Serializer(obj).data)
                            else:
                                result['errors'][model_key].append({
                                    'id': object_id,
                                    'action': action,
                                    'errors': serializer.errors
                                })

                    except Model.DoesNotExist:
                        # Object doesn't exist, try to create
                        serializer = Serializer(data=data)
                        if serializer.is_valid():
                            obj = serializer.save()
                            result['created'][model_key].append(Serializer(obj).data)
                        else:
                            result['errors'][model_key].append({
                                'id': object_id,
                                'action': 'create',
                                'errors': serializer.errors
                            })

                elif action == 'delete':
                    if not object_id:
                        result['errors'][model_key].append({
                            'id': None,
                            'action': action,
                            'errors': {'detail': 'Missing id for delete'}
                        })
                        continue

                    try:
                        obj = Model.objects.get(id=object_id)
                        obj.delete()
                        result['deleted'][model_key].append({'id': object_id})
                    except Model.DoesNotExist:
                        # Already deleted
                        result['deleted'][model_key].append({'id': object_id})

                else:
                    result['errors'][model_key].append({
                        'id': object_id,
                        'action': action,
                        'errors': {'detail': 'Unknown action'}
                    })

            except Exception as e:
                result['errors'][model_key].append({
                    'id': object_id,
                    'action': action,
                    'errors': {'detail': str(e)}
                })

    return Response(result)


# Web Views
@login_required
def dashboard_view(request):
    """
    Dashboard view with optimized queries and caching.
    Shows user-specific dashboard based on role and permissions.
    """
    user = request.user
    context = {
        'user': user,
        'role': user.role,
    }

    try:
        if user.role == 'super_admin':
            # Super admin sees global statistics
            cache_key_schools = 'schools_count'
            cache_key_users = 'users_count'
            context['schools_count'] = cache.get(cache_key_schools, School.objects.count())
            context['users_count'] = cache.get(cache_key_users, User.objects.count())
            # Cache for 5 minutes
            cache.set(cache_key_schools, context['schools_count'], 300)
            cache.set(cache_key_users, context['users_count'], 300)
        elif user.school:
            # Regular users see school-specific statistics
            context['school'] = user.school
            cache_key_classes = f'classes_count_{user.school.id}'
            cache_key_subjects = f'subjects_count_{user.school.id}'
            cache_key_students = f'students_count_{user.school.id}'
            
            # Use select_related for optimized queries
            context['classes_count'] = cache.get(cache_key_classes, 
                ClassSection.objects.filter(school=user.school).count())
            context['subjects_count'] = cache.get(cache_key_subjects, 
                Subject.objects.filter(school=user.school).count())
            context['students_count'] = cache.get(cache_key_students, 
                StudentEnrollment.objects.filter(school=user.school).count())
            
            # Cache for 5 minutes
            cache.set(cache_key_classes, context['classes_count'], 300)
            cache.set(cache_key_subjects, context['subjects_count'], 300)
            cache.set(cache_key_students, context['students_count'], 300)

        return render(request, 'dashboard.html', context)
    
    except Exception as e:
        # Log error and return safe context
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error loading dashboard for user {user.username}: {str(e)}")
        
        # Return minimal context on error
        return render(request, 'dashboard.html', {
            'user': user,
            'role': user.role,
            'error': 'Unable to load dashboard statistics. Please try again later.'
        })


def landing_view(request):
    return render(request, 'landing.html', {
        'title': 'ReportCardApp - Multi-Tenant Report Card System'
    })


def offline_view(request):
    """Offline fallback page for PWA"""
    return render(request, 'offline.html', {
        'title': 'Offline - ReportCardApp'
    })


def manifest_view(request):
    """Serve the web manifest at /manifest.json for PWABuilder and installers."""
    manifest_path = finders.find('manifest.json')
    if not manifest_path:
        manifest_path = os.path.join(settings.BASE_DIR, 'static', 'manifest.json')
    if not os.path.exists(manifest_path):
        raise Http404('Manifest not found')
    return FileResponse(open(manifest_path, 'rb'), content_type='application/manifest+json')


def sw_view(request):
    """Serve the service worker at /sw.js so it is discoverable at web root."""
    sw_path = finders.find('sw.js')
    if not sw_path:
        sw_path = os.path.join(settings.BASE_DIR, 'static', 'sw.js')
    if not os.path.exists(sw_path):
        raise Http404('Service worker not found')
    return FileResponse(open(sw_path, 'rb'), content_type='application/javascript')


def apk_download_view(request):
    """Serve mobile app packages for download."""
    import os
    from django.conf import settings
    from django.http import HttpResponse

    downloads_dir = os.path.join(settings.BASE_DIR, 'downloads')

    # Define available packages
    packages = {
        'android': {
            'filename': 'reportcard.apk',
            'content_type': 'application/vnd.android.package-archive',
            'display_name': 'ReportCardApp.apk'
        },
        'ios': {
            'filename': 'reportcard.ipa',
            'content_type': 'application/octet-stream',
            'display_name': 'ReportCardApp.ipa'
        },
        'windows': {
            'filename': 'reportcard.msix',
            'content_type': 'application/octet-stream',
            'display_name': 'ReportCardApp.msix'
        }
    }

    # Check which package is requested
    package_type = request.GET.get('type', 'android')

    if package_type not in packages:
        return HttpResponse(
            f"Invalid package type '{package_type}'. Available types: {', '.join(packages.keys())}",
            status=400,
            content_type='text/plain'
        )

    package_info = packages[package_type]
    package_path = os.path.join(downloads_dir, package_info['filename'])

    if not os.path.exists(package_path):
        available_packages = []
        for pkg_type, pkg_info in packages.items():
            if os.path.exists(os.path.join(downloads_dir, pkg_info['filename'])):
                available_packages.append(pkg_type)

        if available_packages:
            return HttpResponse(
                f"{package_info['filename']} not found. Available packages: {', '.join(available_packages)}\n"
                f"Download URLs:\n" +
                '\n'.join([f"/download/apk/?type={pkg}" for pkg in available_packages]),
                status=404,
                content_type='text/plain'
            )
        else:
            return HttpResponse(
                "No mobile packages found. Please generate packages using PWABuilder:\n"
                "1. Visit https://www.pwabuilder.com\n"
                "2. Enter your app URL\n"
                "3. Generate and download packages\n"
                "4. Place APK/IPA/MSIX files in the downloads/ directory\n\n"
                "Expected filenames:\n" +
                '\n'.join([f"- downloads/{info['filename']}" for info in packages.values()]),
                status=404,
                content_type='text/plain'
            )

    try:
        response = FileResponse(
            open(package_path, 'rb'),
            content_type=package_info['content_type']
        )
        response['Content-Disposition'] = f'attachment; filename="{package_info["display_name"]}"'
        return response
    except Exception as e:
        return HttpResponse(
            f"Error serving {package_info['filename']}: {str(e)}",
            status=500,
            content_type='text/plain'
        )


@login_required
def school_switch(request):
    if request.user.role != 'super_admin':
        messages.error(request, 'Access denied. Super admin required.')
        return redirect('dashboard')

    # Handle quick switch via GET parameter
    school_id_param = request.GET.get('school_id')
    if school_id_param is not None:
        if school_id_param:
            request.session['school_id'] = int(school_id_param)
            try:
                school = School.objects.get(id=school_id_param)
                messages.success(request, f'Switched to school: {school.name}')
            except School.DoesNotExist:
                messages.error(request, 'School not found')
        else:
            request.session.pop('school_id', None)
            messages.success(request, 'Switched to global view')
        return redirect('dashboard')

    if request.method == 'POST':
        school_id = request.POST.get('school_id')
        if school_id:
            request.session['school_id'] = int(school_id)
            school = School.objects.get(id=school_id)
            messages.success(request, f'Switched to school: {school.name}')
        else:
            request.session.pop('school_id', None)
            messages.success(request, 'Switched to global view')
        return redirect('dashboard')

    schools = School.objects.all()
    current_school_id = request.session.get('school_id')
    current_school = None
    if current_school_id:
        try:
            current_school = School.objects.get(id=current_school_id)
        except School.DoesNotExist:
            pass

    return render(request, 'schools/school_switch.html', {
        'schools': schools,
        'current_school': current_school,
        'title': 'Switch School Context'
    })


# Management Views - Super Admin Only
@login_required
@login_required
def school_list(request):
    if request.user.role != 'super_admin':
        messages.error(request, 'Access denied. Super admin required.')
        return redirect('dashboard')

    schools = School.objects.all().order_by('-created_at')
    return render(request, 'schools/school_list.html', {
        'schools': schools,
        'title': 'Manage Schools'
    })


@login_required
def school_create(request):
    if request.user.role != 'super_admin':
        messages.error(request, 'Access denied. Super admin required.')
        return redirect('dashboard')

    from .forms import SchoolForm
    if request.method == 'POST':
        form = SchoolForm(request.POST)
        if form.is_valid():
            # Extra safety: check duplicates server-side before saving
            name = form.cleaned_data.get('name', '').strip()
            if name and School.objects.filter(name__iexact=name).exists():
                form.add_error('name', 'A school with this name already exists. Please choose a different name.')
                # Display specific field errors to the user
                for field, errors in form.errors.items():
                    for error in errors:
                        if field == '__all__':
                            messages.error(request, str(error))
                        else:
                            messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
                return render(request, 'schools/school_form.html', {
                    'form': form,
                    'title': 'Create School',
                    'action': 'Create'
                })

            try:
                school = form.save()
                messages.success(request, f'School "{school.name}" created successfully.')
                # Redirect with refresh flag so client-side cache is refreshed
                return redirect(reverse('school_list') + '?refresh=1')
            except IntegrityError as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"IntegrityError creating school: {str(e)}", exc_info=True)
                form.add_error('name', 'A school with this name already exists or there was a database constraint error.')
                messages.error(request, 'A school with this name already exists. Please choose a different name.')
                return render(request, 'schools/school_form.html', {
                    'form': form,
                    'title': 'Create School',
                    'action': 'Create'
                })
            except Exception as e:
                # Log the error for debugging
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error creating school: {str(e)}", exc_info=True)
                messages.error(request, 'An error occurred while creating the school. Please try again.')
                # Re-render form with the error
                return render(request, 'schools/school_form.html', {
                    'form': form,
                    'title': 'Create School',
                    'action': 'Create'
                })
        else:
            # Form validation failed - display specific errors
            for field, errors in form.errors.items():
                for error in errors:
                    if field == '__all__':
                        messages.error(request, str(error))
                    else:
                        messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
            # Re-render form with validation errors
            return render(request, 'schools/school_form.html', {
                'form': form,
                'title': 'Create School',
                'action': 'Create'
            })
    else:
        form = SchoolForm()
    
    return render(request, 'schools/school_form.html', {
        'form': form,
        'title': 'Create School',
        'action': 'Create'
    })


@login_required
def school_update(request, pk):
    if request.user.role != 'super_admin':
        messages.error(request, 'Access denied. Super admin required.')
        return redirect('dashboard')

    from .forms import SchoolForm
    school = get_object_or_404(School, pk=pk)
    if request.method == 'POST':
        form = SchoolForm(request.POST, instance=school)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'School updated successfully.')
                return redirect(reverse('school_list') + '?refresh=1')
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error updating school: {str(e)}", exc_info=True)
                messages.error(request, 'An error occurred while updating the school. Please try again.')
        else:
            # Form validation failed - display specific errors
            for field, errors in form.errors.items():
                for error in errors:
                    if field == '__all__':
                        messages.error(request, str(error))
                    else:
                        messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
    else:
        form = SchoolForm(instance=school)
    return render(request, 'schools/school_form.html', {
        'form': form,
        'school': school,
        'title': 'Edit School'
    })


@login_required
@login_required
def school_delete(request, pk):
    if request.user.role != 'super_admin':
        messages.error(request, 'Access denied. Super admin required.')
        return redirect('dashboard')

    school = get_object_or_404(School, pk=pk)
    if request.method == 'POST':
        school.delete()
        messages.success(request, 'School deleted successfully.')
        return redirect(reverse('school_list') + '?refresh=1')
    return render(request, 'schools/school_confirm_delete.html', {
        'school': school,
        'title': 'Delete School'
    })


# User Management Views
@login_required
def user_list(request):
    if request.user.role not in ['super_admin', 'admin']:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboard')

    users = User.objects.all().order_by('-date_joined')
    if request.user.role == 'admin':
        users = users.filter(school=request.user.school)

    # Filter by role if specified
    role_filter = request.GET.get('role')
    if role_filter:
        users = users.filter(role=role_filter)

    return render(request, 'users/user_list.html', {
        'users': users,
        'role_filter': role_filter,
        'title': 'Manage Users'
    })


@login_required
def user_create(request):
    if request.user.role not in ['super_admin', 'admin']:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboard')

    from .forms import UserForm
    if request.method == 'POST':
        form = UserForm(request.POST, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, 'User created successfully.')
            return redirect('user_list')
    else:
        form = UserForm(request=request)
    return render(request, 'users/user_form.html', {
        'form': form,
        'title': 'Create User'
    })


@login_required
def user_update(request, pk):
    if request.user.role not in ['super_admin', 'admin']:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboard')

    from .forms import UserForm
    user_obj = get_object_or_404(User, pk=pk)

    # Check if admin can only edit users from their school
    if request.user.role == 'admin' and user_obj.school != request.user.school:
        messages.error(request, 'Access denied. Cannot edit users from other schools.')
        return redirect('user_list')

    if request.method == 'POST':
        form = UserForm(request.POST, instance=user_obj, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, 'User updated successfully.')
            return redirect('user_list')
    else:
        form = UserForm(instance=user_obj, request=request)
    return render(request, 'users/user_form.html', {
        'form': form,
        'user_obj': user_obj,
        'title': 'Edit User'
    })


@login_required
def user_delete(request, pk):
    if request.user.role not in ['super_admin', 'admin']:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboard')

    user_obj = get_object_or_404(User, pk=pk)

    # Check if admin can only delete users from their school
    if request.user.role == 'admin' and user_obj.school != request.user.school:
        messages.error(request, 'Access denied. Cannot delete users from other schools.')
        return redirect('user_list')

    if request.method == 'POST':
        user_obj.delete()
        messages.success(request, 'User deleted successfully.')
        return redirect('user_list')
    return render(request, 'users/user_confirm_delete.html', {
        'user_obj': user_obj,
        'title': 'Delete User'
    })


# Academic Management Views
@login_required
def class_section_list(request):
    if request.user.role not in ['super_admin', 'admin']:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboard')

    class_sections = ClassSection.objects.all().order_by('school', 'name')
    if request.user.role == 'admin':
        class_sections = class_sections.filter(school=request.user.school)

    return render(request, 'class_sections/class_section_list.html', {
        'class_sections': class_sections,
        'title': 'Manage Class Sections'
    })


@login_required
def class_section_create(request):
    if request.user.role not in ['super_admin', 'admin']:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboard')

    from .forms import ClassSectionForm
    if request.method == 'POST':
        form = ClassSectionForm(request.POST, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, 'Class section created successfully.')
            return redirect('class_section_list')
    else:
        form = ClassSectionForm(request=request)
    return render(request, 'class_sections/class_section_form.html', {
        'form': form,
        'title': 'Create Class Section'
    })


@login_required
def class_section_update(request, pk):
    if request.user.role not in ['super_admin', 'admin']:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboard')

    from .forms import ClassSectionForm
    class_section = get_object_or_404(ClassSection, pk=pk)

    # Check if admin can only edit class sections from their school
    if request.user.role == 'admin' and class_section.school != request.user.school:
        messages.error(request, 'Access denied. Cannot edit class sections from other schools.')
        return redirect('class_section_list')

    if request.method == 'POST':
        form = ClassSectionForm(request.POST, instance=class_section, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, 'Class section updated successfully.')
            return redirect('class_section_list')
    else:
        form = ClassSectionForm(instance=class_section, request=request)
    return render(request, 'class_sections/class_section_form.html', {
        'form': form,
        'class_section': class_section,
        'title': 'Edit Class Section'
    })


@login_required
def class_section_delete(request, pk):
    if request.user.role not in ['super_admin', 'admin']:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboard')

    class_section = get_object_or_404(ClassSection, pk=pk)

    # Check if admin can only delete class sections from their school
    if request.user.role == 'admin' and class_section.school != request.user.school:
        messages.error(request, 'Access denied. Cannot delete class sections from other schools.')
        return redirect('class_section_list')

    if request.method == 'POST':
        class_section.delete()
        messages.success(request, 'Class section deleted successfully.')
        return redirect('class_section_list')
    return render(request, 'class_sections/class_section_confirm_delete.html', {
        'class_section': class_section,
        'title': 'Delete Class Section'
    })


# Subject Management Views
@login_required
def subject_list(request):
    if request.user.role not in ['super_admin', 'admin']:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboard')

    subjects = Subject.objects.all().order_by('school', 'name')
    if request.user.role == 'admin':
        subjects = subjects.filter(school=request.user.school)

    return render(request, 'subjects/subject_list.html', {
        'subjects': subjects,
        'title': 'Manage Subjects'
    })


@login_required
def subject_create(request):
    if request.user.role not in ['super_admin', 'admin']:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboard')

    from .forms import SubjectForm
    if request.method == 'POST':
        form = SubjectForm(request.POST, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, 'Subject created successfully.')
            return redirect('subject_list')
    else:
        form = SubjectForm(request=request)
    return render(request, 'subjects/subject_form.html', {
        'form': form,
        'title': 'Create Subject'
    })


@login_required
def subject_update(request, pk):
    if request.user.role not in ['super_admin', 'admin']:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboard')

    from .forms import SubjectForm
    subject = get_object_or_404(Subject, pk=pk)

    # Check if admin can only edit subjects from their school
    if request.user.role == 'admin' and subject.school != request.user.school:
        messages.error(request, 'Access denied. Cannot edit subjects from other schools.')
        return redirect('subject_list')

    if request.method == 'POST':
        form = SubjectForm(request.POST, instance=subject, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, 'Subject updated successfully.')
            return redirect('subject_list')
    else:
        form = SubjectForm(instance=subject, request=request)
    return render(request, 'subjects/subject_form.html', {
        'form': form,
        'subject': subject,
        'title': 'Edit Subject'
    })


@login_required
def subject_delete(request, pk):
    if request.user.role not in ['super_admin', 'admin']:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboard')

    subject = get_object_or_404(Subject, pk=pk)

    # Check if admin can only delete subjects from their school
    if request.user.role == 'admin' and subject.school != request.user.school:
        messages.error(request, 'Access denied. Cannot delete subjects from other schools.')
        return redirect('subject_list')

    if request.method == 'POST':
        subject.delete()
        messages.success(request, 'Subject deleted successfully.')
        return redirect('subject_list')
    return render(request, 'subjects/subject_confirm_delete.html', {
        'subject': subject,
        'title': 'Delete Subject'
    })


# Grading Scale Management Views
@login_required
def grading_scale_list(request):
    if request.user.role not in ['super_admin', 'admin']:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboard')

    grading_scales = GradingScale.objects.all().order_by('school', 'name')
    if request.user.role == 'admin':
        grading_scales = grading_scales.filter(school=request.user.school)

    return render(request, 'grading_scales/grading_scale_list.html', {
        'grading_scales': grading_scales,
        'title': 'Manage Grading Scales'
    })


@login_required
def grading_scale_create(request):
    if request.user.role not in ['super_admin', 'admin']:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboard')

    from .forms import GradingScaleForm
    if request.method == 'POST':
        form = GradingScaleForm(request.POST, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, 'Grading scale created successfully.')
            return redirect('grading_scale_list')
    else:
        form = GradingScaleForm(request=request)
    return render(request, 'grading_scales/grading_scale_form.html', {
        'form': form,
        'title': 'Create Grading Scale'
    })


@login_required
def grading_scale_update(request, pk):
    if request.user.role not in ['super_admin', 'admin']:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboard')

    from .forms import GradingScaleForm
    grading_scale = get_object_or_404(GradingScale, pk=pk)

    # Check if admin can only edit grading scales from their school
    if request.user.role == 'admin' and grading_scale.school != request.user.school:
        messages.error(request, 'Access denied. Cannot edit grading scales from other schools.')
        return redirect('grading_scale_list')

    if request.method == 'POST':
        form = GradingScaleForm(request.POST, instance=grading_scale, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, 'Grading scale updated successfully.')
            return redirect('grading_scale_list')
    else:
        form = GradingScaleForm(instance=grading_scale, request=request)
    return render(request, 'grading_scales/grading_scale_form.html', {
        'form': form,
        'grading_scale': grading_scale,
        'title': 'Edit Grading Scale'
    })


@login_required
def grading_scale_delete(request, pk):
    if request.user.role not in ['super_admin', 'admin']:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboard')

    grading_scale = get_object_or_404(GradingScale, pk=pk)

    # Check if admin can only delete grading scales from their school
    if request.user.role == 'admin' and grading_scale.school != request.user.school:
        messages.error(request, 'Access denied. Cannot delete grading scales from other schools.')
        return redirect('grading_scale_list')

    if request.method == 'POST':
        grading_scale.delete()
        messages.success(request, 'Grading scale deleted successfully.')
        return redirect('grading_scale_list')
    return render(request, 'grading_scales/grading_scale_confirm_delete.html', {
        'grading_scale': grading_scale,
        'title': 'Delete Grading Scale'
    })


# Student Enrollment Management Views
@login_required
def enrollment_list(request):
    if request.user.role not in ['super_admin', 'admin']:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboard')

    enrollments = StudentEnrollment.objects.all().select_related('student', 'class_section', 'school').order_by('school', 'class_section', 'student__last_name')
    if request.user.role == 'admin':
        enrollments = enrollments.filter(school=request.user.school)

    return render(request, 'enrollments/enrollment_list.html', {
        'enrollments': enrollments,
        'title': 'Manage Student Enrollments'
    })


@login_required
def enrollment_create(request):
    if request.user.role not in ['super_admin', 'admin']:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboard')

    from .forms import StudentEnrollmentForm
    if request.method == 'POST':
        form = StudentEnrollmentForm(request.POST, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, 'Student enrollment created successfully.')
            return redirect('enrollment_list')
    else:
        form = StudentEnrollmentForm(request=request)
    return render(request, 'enrollments/enrollment_form.html', {
        'form': form,
        'title': 'Create Student Enrollment'
    })


@login_required
def enrollment_update(request, pk):
    if request.user.role not in ['super_admin', 'admin']:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboard')

    from .forms import StudentEnrollmentForm
    enrollment = get_object_or_404(StudentEnrollment, pk=pk)

    # Check if admin can only edit enrollments from their school
    if request.user.role == 'admin' and enrollment.school != request.user.school:
        messages.error(request, 'Access denied. Cannot edit enrollments from other schools.')
        return redirect('enrollment_list')

    if request.method == 'POST':
        form = StudentEnrollmentForm(request.POST, instance=enrollment, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, 'Student enrollment updated successfully.')
            return redirect('enrollment_list')
    else:
        form = StudentEnrollmentForm(instance=enrollment, request=request)
    return render(request, 'enrollments/enrollment_form.html', {
        'form': form,
        'enrollment': enrollment,
        'title': 'Edit Student Enrollment'
    })


@login_required
def enrollment_delete(request, pk):
    if request.user.role not in ['super_admin', 'admin']:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboard')

    enrollment = get_object_or_404(StudentEnrollment, pk=pk)

    # Check if admin can only delete enrollments from their school
    if request.user.role == 'admin' and enrollment.school != request.user.school:
        messages.error(request, 'Access denied. Cannot delete enrollments from other schools.')
        return redirect('enrollment_list')

    if request.method == 'POST':
        enrollment.delete()
        messages.success(request, 'Student enrollment deleted successfully.')
        return redirect('enrollment_list')
    return render(request, 'enrollments/enrollment_confirm_delete.html', {
        'enrollment': enrollment,
        'title': 'Delete Student Enrollment'
    })


# Grading Period Management Views
@login_required
def grading_period_list(request):
    if request.user.role not in ['super_admin', 'admin']:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboard')

    grading_periods = GradingPeriod.objects.all().order_by('school', 'start_date')
    if request.user.role == 'admin':
        grading_periods = grading_periods.filter(school=request.user.school)

    return render(request, 'grading_periods/grading_period_list.html', {
        'grading_periods': grading_periods,
        'title': 'Manage Grading Periods'
    })


@login_required
def grading_period_create(request):
    if request.user.role not in ['super_admin', 'admin']:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboard')

    from .forms import GradingPeriodForm
    if request.method == 'POST':
        form = GradingPeriodForm(request.POST, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, 'Grading period created successfully.')
            return redirect('grading_period_list')
    else:
        form = GradingPeriodForm(request=request)
    return render(request, 'grading_periods/grading_period_form.html', {
        'form': form,
        'title': 'Create Grading Period'
    })


@login_required
def grading_period_update(request, pk):
    if request.user.role not in ['super_admin', 'admin']:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboard')

    from .forms import GradingPeriodForm
    grading_period = get_object_or_404(GradingPeriod, pk=pk)

    # Check if admin can only edit grading periods from their school
    if request.user.role == 'admin' and grading_period.school != request.user.school:
        messages.error(request, 'Access denied. Cannot edit grading periods from other schools.')
        return redirect('grading_period_list')

    if request.method == 'POST':
        form = GradingPeriodForm(request.POST, instance=grading_period, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, 'Grading period updated successfully.')
            return redirect('grading_period_list')
    else:
        form = GradingPeriodForm(instance=grading_period, request=request)
    return render(request, 'grading_periods/grading_period_form.html', {
        'form': form,
        'grading_period': grading_period,
        'title': 'Edit Grading Period'
    })


@login_required
def grading_period_delete(request, pk):
    if request.user.role not in ['super_admin', 'admin']:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboard')

    grading_period = get_object_or_404(GradingPeriod, pk=pk)

    # Check if admin can only delete grading periods from their school
    if request.user.role == 'admin' and grading_period.school != request.user.school:
        messages.error(request, 'Access denied. Cannot delete grading periods from other schools.')
        return redirect('grading_period_list')

    if request.method == 'POST':
        grading_period.delete()
        messages.success(request, 'Grading period deleted successfully.')
        return redirect('grading_period_list')
    return render(request, 'grading_periods/grading_period_confirm_delete.html', {
        'grading_period': grading_period,
        'title': 'Delete Grading Period'
    })


# Grade Management Views
@login_required
def grade_list(request):
    if request.user.role not in ['super_admin', 'admin', 'teacher']:
        messages.error(request, 'Access denied. Insufficient privileges.')
        return redirect('dashboard')

    grades = Grade.objects.all().select_related('student', 'subject', 'grading_period', 'school').order_by('school', 'grading_period', 'subject', 'student__last_name')
    if request.user.role == 'admin':
        grades = grades.filter(school=request.user.school)
    elif request.user.role == 'teacher':
        grades = grades.filter(school=request.user.school, subject__class_sections__teacher=request.user).distinct()

    # Filter by grading period if specified
    grading_period_id = request.GET.get('grading_period')
    if grading_period_id:
        grades = grades.filter(grading_period_id=grading_period_id)

    return render(request, 'grades/grade_list.html', {
        'grades': grades,
        'grading_period_id': grading_period_id,
        'title': 'Manage Grades'
    })


@login_required
def grade_bulk_entry(request):
    if request.user.role not in ['super_admin', 'admin', 'teacher']:
        messages.error(request, 'Access denied. Insufficient privileges.')
        return redirect('dashboard')

    school = request.user.school if request.user.role != 'super_admin' else None

    # Get available subjects for the teacher/admin
    subjects = Subject.objects.filter(school=school) if school else Subject.objects.all()
    if request.user.role == 'teacher':
        subjects = subjects.filter(class_sections__teacher=request.user).distinct()

    # Get available grading periods
    grading_periods = GradingPeriod.objects.filter(school=school) if school else GradingPeriod.objects.all()

    selected_subject_id = request.GET.get('subject')
    selected_grading_period_id = request.GET.get('grading_period')
    selected_class_id = request.GET.get('class_section')

    students = []
    existing_grades = {}

    if selected_subject_id and selected_grading_period_id and selected_class_id:
        try:
            subject = Subject.objects.get(id=selected_subject_id, school=school) if school else Subject.objects.get(id=selected_subject_id)
            grading_period = GradingPeriod.objects.get(id=selected_grading_period_id, school=school) if school else GradingPeriod.objects.get(id=selected_grading_period_id)
            class_section = ClassSection.objects.get(id=selected_class_id, school=school) if school else ClassSection.objects.get(id=selected_class_id)

            # Get enrolled students for this class
            enrollments = StudentEnrollment.objects.filter(class_section=class_section).select_related('student')
            students = [enrollment.student for enrollment in enrollments]

            # Get existing grades for this subject/grading period
            grades = Grade.objects.filter(
                student__in=students,
                subject=subject,
                grading_period=grading_period
            ).select_related('student')

            existing_grades = {grade.student.id: grade for grade in grades}

        except (Subject.DoesNotExist, GradingPeriod.DoesNotExist, ClassSection.DoesNotExist):
            messages.error(request, 'Invalid selection.')
            return redirect('grade_bulk_entry')

    # Handle POST request for bulk grade submission
    if request.method == 'POST':
        for key, value in request.POST.items():
            if key.startswith('score_'):
                student_id = key.split('_')[1]
                score = value.strip()
                comments = request.POST.get(f'comments_{student_id}', '').strip()

                if score:
                    try:
                        score_float = float(score)
                        student = User.objects.get(id=student_id, role='student')

                        # Get or create grade
                        grade, created = Grade.objects.get_or_create(
                            student=student,
                            subject_id=selected_subject_id,
                            grading_period_id=selected_grading_period_id,
                            defaults={'school': school or student.school}
                        )

                        grade.score = score_float
                        grade.comments = comments
                        grade.save()

                    except (ValueError, User.DoesNotExist):
                        continue

        messages.success(request, 'Grades saved successfully.')
        return redirect('grade_bulk_entry')

    # Get available class sections for the selected subject
    class_sections = []
    if selected_subject_id:
        class_sections = ClassSection.objects.filter(
            school=school,
            subjects__id=selected_subject_id
        ).distinct() if school else ClassSection.objects.filter(subjects__id=selected_subject_id).distinct()

    context = {
        'subjects': subjects,
        'grading_periods': grading_periods,
        'class_sections': class_sections,
        'students': students,
        'existing_grades': existing_grades,
        'selected_subject_id': selected_subject_id,
        'selected_grading_period_id': selected_grading_period_id,
        'selected_class_id': selected_class_id,
        'title': 'Bulk Grade Entry'
    }

    return render(request, 'grades/grade_bulk_entry.html', context)


@login_required
def grade_import(request):
    if request.user.role not in ['super_admin', 'admin', 'teacher']:
        messages.error(request, 'Access denied. Insufficient privileges.')
        return redirect('dashboard')

    school = request.user.school if request.user.role != 'super_admin' else None

    if request.method == 'POST':
        import_file = request.FILES.get('import_file')
        if not import_file:
            messages.error(request, 'Please select a file to import.')
            return redirect('grade_import')

        # Process file
        try:
            if import_file.name.endswith('.xlsx') or import_file.name.endswith('.xls'):
                # Excel file processing
                try:
                    import pandas as pd
                except ImportError:
                    messages.error(request, 'Pandas library is required for Excel file processing. Please install pandas.')
                    return redirect('grade_import')

                df = pd.read_excel(import_file)

                # Expected columns: student_id, subject_code, grading_period_name, score, comments
                required_columns = ['student_id', 'subject_code', 'grading_period_name', 'score']
                if not all(col in df.columns for col in required_columns):
                    messages.error(request, f'File must contain columns: {", ".join(required_columns)}')
                    return redirect('grade_import')

                success_count = 0
                error_count = 0

                for _, row in df.iterrows():
                    try:
                        student = User.objects.get(username=row['student_id'], role='student')
                        if school and student.school != school:
                            error_count += 1
                            continue

                        subject = Subject.objects.get(code=row['subject_code'])
                        if school and subject.school != school:
                            error_count += 1
                            continue

                        grading_period = GradingPeriod.objects.get(name=row['grading_period_name'])
                        if school and grading_period.school != school:
                            error_count += 1
                            continue

                        # Create or update grade
                        grade, created = Grade.objects.get_or_create(
                            student=student,
                            subject=subject,
                            grading_period=grading_period,
                            defaults={'school': school or student.school}
                        )

                        grade.score = float(row['score'])
                        grade.comments = row.get('comments', '') if pd.notna(row.get('comments')) else ''
                        grade.save()

                        success_count += 1

                    except (User.DoesNotExist, Subject.DoesNotExist, GradingPeriod.DoesNotExist, ValueError) as e:
                        error_count += 1
                        continue

                messages.success(request, f'Import completed. {success_count} grades imported successfully, {error_count} errors.')

            elif import_file.name.endswith('.csv'):
                # CSV file processing
                import csv
                import io

                # Read CSV content
                file_data = import_file.read().decode('utf-8')
                csv_reader = csv.DictReader(io.StringIO(file_data))

                success_count = 0
                error_count = 0

                for row in csv_reader:
                    try:
                        student = User.objects.get(username=row['student_id'], role='student')
                        if school and student.school != school:
                            error_count += 1
                            continue

                        subject = Subject.objects.get(code=row['subject_code'])
                        if school and subject.school != school:
                            error_count += 1
                            continue

                        grading_period = GradingPeriod.objects.get(name=row['grading_period_name'])
                        if school and grading_period.school != school:
                            error_count += 1
                            continue

                        # Create or update grade
                        grade, created = Grade.objects.get_or_create(
                            student=student,
                            subject=subject,
                            grading_period=grading_period,
                            defaults={'school': school or student.school}
                        )

                        grade.score = float(row['score'])
                        grade.comments = row.get('comments', '')
                        grade.save()

                        success_count += 1

                    except (User.DoesNotExist, Subject.DoesNotExist, GradingPeriod.DoesNotExist, ValueError, KeyError) as e:
                        error_count += 1
                        continue

                messages.success(request, f'Import completed. {success_count} grades imported successfully, {error_count} errors.')

            else:
                messages.error(request, 'Unsupported file format. Please use Excel (.xlsx, .xls) or CSV (.csv) files.')

        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')

        return redirect('grade_import')

    return render(request, 'grades/grade_import.html', {
        'title': 'Import Grades from Excel/CSV'
    })


@login_required
def grade_create(request):
    if request.user.role not in ['super_admin', 'admin', 'teacher']:
        messages.error(request, 'Access denied. Insufficient privileges.')
        return redirect('dashboard')

    from .forms import GradeForm
    if request.method == 'POST':
        form = GradeForm(request.POST, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, 'Grade created successfully.')
            return redirect('grade_list')
    else:
        form = GradeForm(request=request)
    return render(request, 'grades/grade_form.html', {
        'form': form,
        'title': 'Create Grade'
    })


@login_required
def grade_update(request, pk):
    if request.user.role not in ['super_admin', 'admin', 'teacher']:
        messages.error(request, 'Access denied. Insufficient privileges.')
        return redirect('dashboard')

    from .forms import GradeForm
    grade = get_object_or_404(Grade, pk=pk)

    # Enhanced permissions check with proper authorization
    if request.user.role == 'super_admin':
        # Super admin can edit any grade
        pass
    elif request.user.role == 'admin':
        # Admin can only edit grades from their school
        if grade.school != request.user.school:
            messages.error(request, 'Access denied. Cannot edit grades from other schools.')
            return redirect('grade_list')
    elif request.user.role == 'teacher':
        # Teacher can only edit grades for subjects they teach in their school
        if grade.school != request.user.school:
            messages.error(request, 'Access denied. Cannot edit grades from other schools.')
            return redirect('grade_list')
        
        # Check if teacher teaches this subject
        if not grade.subject.class_sections.filter(teacher=request.user).exists():
            messages.error(request, 'Access denied. Cannot edit grades for subjects you do not teach.')
            return redirect('grade_list')
        
        # Additional check: ensure teacher has access to this student's class
        student_in_teacher_class = StudentEnrollment.objects.filter(
            student=grade.student,
            class_section__teacher=request.user
        ).exists()
        
        if not student_in_teacher_class:
            messages.error(request, 'Access denied. Cannot edit grades for students not in your classes.')
            return redirect('grade_list')
    else:
        messages.error(request, 'Access denied. Insufficient privileges.')
        return redirect('grade_list')

    if request.method == 'POST':
        form = GradeForm(request.POST, instance=grade, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, 'Grade updated successfully.')
            return redirect('grade_list')
    else:
        form = GradeForm(instance=grade, request=request)
    return render(request, 'grades/grade_form.html', {
        'form': form,
        'grade': grade,
        'title': 'Edit Grade'
    })


@login_required
def grade_delete(request, pk):
    if request.user.role not in ['super_admin', 'admin', 'teacher']:
        messages.error(request, 'Access denied. Insufficient privileges.')
        return redirect('dashboard')

    grade = get_object_or_404(Grade, pk=pk)

    # Enhanced permissions check with proper authorization
    if request.user.role == 'super_admin':
        # Super admin can delete any grade
        pass
    elif request.user.role == 'admin':
        # Admin can only delete grades from their school
        if grade.school != request.user.school:
            messages.error(request, 'Access denied. Cannot delete grades from other schools.')
            return redirect('grade_list')
    elif request.user.role == 'teacher':
        # Teacher can only delete grades for subjects they teach in their school
        if grade.school != request.user.school:
            messages.error(request, 'Access denied. Cannot delete grades from other schools.')
            return redirect('grade_list')
        
        # Check if teacher teaches this subject
        if not grade.subject.class_sections.filter(teacher=request.user).exists():
            messages.error(request, 'Access denied. Cannot delete grades for subjects you do not teach.')
            return redirect('grade_list')
        
        # Additional check: ensure teacher has access to this student's class
        student_in_teacher_class = StudentEnrollment.objects.filter(
            student=grade.student,
            class_section__teacher=request.user
        ).exists()
        
        if not student_in_teacher_class:
            messages.error(request, 'Access denied. Cannot delete grades for students not in your classes.')
            return redirect('grade_list')
    else:
        messages.error(request, 'Access denied. Insufficient privileges.')
        return redirect('grade_list')

    if request.method == 'POST':
        grade.delete()
        messages.success(request, 'Grade deleted successfully.')
        return redirect('grade_list')
    return render(request, 'grades/grade_confirm_delete.html', {
        'grade': grade,
        'title': 'Delete Grade'
    })


# Attendance Management Views
@login_required
def attendance_list(request):
    if request.user.role not in ['super_admin', 'admin', 'teacher']:
        messages.error(request, 'Access denied. Insufficient privileges.')
        return redirect('dashboard')

    attendances = Attendance.objects.all().select_related('student', 'class_section', 'school').order_by('school', 'date', 'class_section', 'student__last_name')
    if request.user.role == 'admin':
        attendances = attendances.filter(school=request.user.school)
    elif request.user.role == 'teacher':
        attendances = attendances.filter(school=request.user.school, class_section__teacher=request.user)

    # Filter by date if specified
    date_filter = request.GET.get('date')
    if date_filter:
        attendances = attendances.filter(date=date_filter)

    return render(request, 'attendance/attendance_list.html', {
        'attendances': attendances,
        'date_filter': date_filter,
        'title': 'Manage Attendance'
    })


@login_required
def attendance_create(request):
    if request.user.role not in ['super_admin', 'admin', 'teacher']:
        messages.error(request, 'Access denied. Insufficient privileges.')
        return redirect('dashboard')

    from .forms import AttendanceForm
    if request.method == 'POST':
        form = AttendanceForm(request.POST, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, 'Attendance record created successfully.')
            return redirect('attendance_list')
    else:
        form = AttendanceForm(request=request)
    return render(request, 'attendance/attendance_form.html', {
        'form': form,
        'title': 'Create Attendance Record'
    })


@login_required
def attendance_update(request, pk):
    if request.user.role not in ['super_admin', 'admin', 'teacher']:
        messages.error(request, 'Access denied. Insufficient privileges.')
        return redirect('dashboard')

    from .forms import AttendanceForm
    attendance = get_object_or_404(Attendance, pk=pk)

    # Check permissions
    if request.user.role == 'admin' and attendance.school != request.user.school:
        messages.error(request, 'Access denied. Cannot edit attendance from other schools.')
        return redirect('attendance_list')
    elif request.user.role == 'teacher' and attendance.class_section.teacher != request.user:
        messages.error(request, 'Access denied. Cannot edit attendance for classes you do not teach.')
        return redirect('attendance_list')

    if request.method == 'POST':
        form = AttendanceForm(request.POST, instance=attendance, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, 'Attendance record updated successfully.')
            return redirect('attendance_list')
    else:
        form = AttendanceForm(instance=attendance, request=request)
    return render(request, 'attendance/attendance_form.html', {
        'form': form,
        'attendance': attendance,
        'title': 'Edit Attendance Record'
    })


@login_required
def attendance_delete(request, pk):
    if request.user.role not in ['super_admin', 'admin', 'teacher']:
        messages.error(request, 'Access denied. Insufficient privileges.')
        return redirect('dashboard')

    attendance = get_object_or_404(Attendance, pk=pk)

    # Check permissions
    if request.user.role == 'admin' and attendance.school != request.user.school:
        messages.error(request, 'Access denied. Cannot delete attendance from other schools.')
        return redirect('attendance_list')
    elif request.user.role == 'teacher' and attendance.class_section.teacher != request.user:
        messages.error(request, 'Access denied. Cannot delete attendance for classes you do not teach.')
        return redirect('attendance_list')

    if request.method == 'POST':
        attendance.delete()
        messages.success(request, 'Attendance record deleted successfully.')
        return redirect('attendance_list')
    return render(request, 'attendance/attendance_confirm_delete.html', {
        'attendance': attendance,
        'title': 'Delete Attendance Record'
    })


# Application Management Views
@login_required
def application_list(request):
    if request.user.role not in ['super_admin', 'admin']:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboard')

    applications = UserApplication.objects.all().select_related('school', 'submitted_by', 'reviewed_by')

    # Filter applications based on user role
    if request.user.role == 'super_admin':
        # Super admin can see admin applications globally and teacher applications for all schools
        applications = applications.filter(role__in=['admin', 'teacher'])
    elif request.user.role == 'admin':
        # School admin can only see teacher applications for their school
        applications = applications.filter(role='teacher', school=request.user.school)

    # Filter by status if specified
    status_filter = request.GET.get('status')
    if status_filter:
        applications = applications.filter(status=status_filter)

    return render(request, 'applications/application_list.html', {
        'applications': applications,
        'status_filter': status_filter,
        'title': 'Manage Applications'
    })


@login_required
def application_review(request, pk):
    application = get_object_or_404(UserApplication, pk=pk)

    # Check permissions
    if request.user.role == 'super_admin':
        # Super admin can review admin and teacher applications
        if application.role not in ['admin', 'teacher']:
            messages.error(request, 'Access denied.')
            return redirect('application_list')
    elif request.user.role == 'admin':
        # School admin can only review teacher applications for their school
        if application.role != 'teacher' or application.school != request.user.school:
            messages.error(request, 'Access denied.')
            return redirect('application_list')
    else:
        messages.error(request, 'Access denied.')
        return redirect('application_list')

    if application.status != 'pending':
        messages.error(request, 'This application has already been reviewed.')
        return redirect('application_list')

    from .forms import ApplicationReviewForm
    if request.method == 'POST':
        form = ApplicationReviewForm(request.POST, request=request)
        if form.is_valid():
            action = form.cleaned_data['action']
            review_notes = form.cleaned_data['review_notes']

            if action == 'approve':
                user = application.approve(request.user)
                if user:
                    messages.success(request, f'Application approved. User {user.username} has been created.')
                else:
                    messages.error(request, 'Failed to approve application.')
            elif action == 'reject':
                if application.reject(request.user, review_notes):
                    messages.success(request, 'Application rejected.')
                else:
                    messages.error(request, 'Failed to reject application.')

            return redirect('application_list')
    else:
        form = ApplicationReviewForm(request=request)

    return render(request, 'applications/application_review.html', {
        'application': application,
        'form': form,
        'title': 'Review Application'
    })


# PDF Generation and Report Card Views
@login_required
def report_card_pdf(request, student_id):
    if request.user.role not in ['super_admin', 'admin', 'teacher']:
        messages.error(request, 'Access denied. Insufficient privileges.')
        return redirect('dashboard')

    student = get_object_or_404(User, id=student_id, role='student')

    # Check permissions
    if request.user.role == 'admin' and student.school != request.user.school:
        messages.error(request, 'Access denied. Cannot view reports for students from other schools.')
        return redirect('dashboard')
    elif request.user.role == 'teacher' and not StudentEnrollment.objects.filter(
        student=student,
        class_section__teacher=request.user
    ).exists():
        messages.error(request, 'Access denied. Cannot view reports for students you do not teach.')
        return redirect('dashboard')

    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from django.http import HttpResponse
    from django.conf import settings
    import io
    import os

    # Create PDF buffer
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    story = []

    # Get student data
    enrollment = StudentEnrollment.objects.filter(student=student).first()
    grades = Grade.objects.filter(student=student).select_related('subject', 'grading_period').order_by('grading_period__start_date', 'subject__name')

    # Get school profile for branding
    try:
        school_profile = SchoolProfile.objects.get(school=student.school)
        report_header = school_profile.report_header or f"{student.school.name}"
        report_footer = school_profile.report_footer or "School Administration"
        report_signature = school_profile.report_signature or "Authorized by Principal"
    except SchoolProfile.DoesNotExist:
        report_header = student.school.name
        report_footer = "School Administration"
        report_signature = "Authorized by Principal"

    # Header with school branding
    story.append(Paragraph(f"<b>{report_header}</b>", styles['Title']))
    story.append(Paragraph("<b>Report Card</b>", styles['Heading1']))
    story.append(Spacer(1, 12))

    # Student info
    story.append(Paragraph(f"<b>Student Name:</b> {student.get_full_name()}", styles['Normal']))
    story.append(Paragraph(f"<b>Student ID:</b> {student.username}", styles['Normal']))
    if enrollment:
        story.append(Paragraph(f"<b>Class:</b> {enrollment.class_section.name}", styles['Normal']))
    story.append(Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%B %d, %Y')}", styles['Normal']))
    story.append(Spacer(1, 20))

    # Grades table
    if grades:
        data = [['Subject', 'Grading Period', 'Score', 'Grade', 'Comments']]
        for grade in grades:
            data.append([
                grade.subject.name,
                grade.grading_period.name,
                str(grade.score) if grade.score else '-',
                grade.letter_grade or '-',
                grade.comments or '-'
            ])

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(table)
    else:
        story.append(Paragraph("No grades available.", styles['Normal']))

    story.append(Spacer(1, 30))
    story.append(Paragraph(f"<i>{report_footer}</i>", styles['Italic']))
    story.append(Spacer(1, 10))
    story.append(Paragraph(f"<i>{report_signature}</i>", styles['Italic']))

    doc.build(story)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{student.username}_report_card.pdf"'

    return response


@login_required
def batch_report_card_pdf(request, class_id):
    """
    Generate batch PDF report cards for all students in a class section.
    """
    if request.user.role not in ['super_admin', 'admin', 'teacher']:
        messages.error(request, 'Access denied. Insufficient privileges.')
        return redirect('dashboard')

    class_section = get_object_or_404(ClassSection, id=class_id)

    # Check permissions
    if request.user.role == 'admin' and class_section.school != request.user.school:
        messages.error(request, 'Access denied. Cannot generate reports for classes from other schools.')
        return redirect('dashboard')
    elif request.user.role == 'teacher' and class_section.teacher != request.user:
        messages.error(request, 'Access denied. Cannot generate reports for classes you do not teach.')
        return redirect('dashboard')

    # Get enrolled students
    enrollments = StudentEnrollment.objects.filter(class_section=class_section).select_related('student')
    students = [enrollment.student for enrollment in enrollments]

    if not students:
        messages.error(request, 'No students enrolled in this class.')
        return redirect('report_card_list')

    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from django.http import HttpResponse
    from django.conf import settings
    import io
    import os

    # Create PDF buffer
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    story = []

    # Get school profile for branding
    try:
        school_profile = SchoolProfile.objects.get(school=class_section.school)
        report_header = school_profile.report_header or f"{class_section.school.name}"
        report_footer = school_profile.report_footer or "School Administration"
        report_signature = school_profile.report_signature or "Authorized by Principal"
    except SchoolProfile.DoesNotExist:
        report_header = class_section.school.name
        report_footer = "School Administration"
        report_signature = "Authorized by Principal"

    first_student = True
    for student in students:
        if not first_student:
            story.append(PageBreak())

        # Get student data
        enrollment = StudentEnrollment.objects.filter(student=student, class_section=class_section).first()
        grades = Grade.objects.filter(student=student).select_related('subject', 'grading_period').order_by('grading_period__start_date', 'subject__name')

        # Header with school branding
        story.append(Paragraph(f"<b>{report_header}</b>", styles['Title']))
        story.append(Paragraph("<b>Report Card</b>", styles['Heading1']))
        story.append(Spacer(1, 12))

        # Student info
        story.append(Paragraph(f"<b>Student Name:</b> {student.get_full_name()}", styles['Normal']))
        story.append(Paragraph(f"<b>Student ID:</b> {student.username}", styles['Normal']))
        story.append(Paragraph(f"<b>Class:</b> {class_section.name}", styles['Normal']))
        story.append(Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%B %d, %Y')}", styles['Normal']))
        story.append(Spacer(1, 20))

        # Grades table
        if grades:
            data = [['Subject', 'Grading Period', 'Score', 'Grade', 'Comments']]
            for grade in grades:
                data.append([
                    grade.subject.name,
                    grade.grading_period.name,
                    str(grade.score) if grade.score else '-',
                    grade.letter_grade or '-',
                    grade.comments or '-'
                ])

            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
        else:
            story.append(Paragraph("No grades available.", styles['Normal']))

        story.append(Spacer(1, 30))
        story.append(Paragraph(f"<i>{report_footer}</i>", styles['Italic']))
        story.append(Spacer(1, 10))
        story.append(Paragraph(f"<i>{report_signature}</i>", styles['Italic']))

        first_student = False

    doc.build(story)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{class_section.name}_report_cards.pdf"'

    return response


@login_required
def report_card_list(request):
    if request.user.role not in ['super_admin', 'admin', 'teacher', 'student']:
        messages.error(request, 'Access denied. Insufficient privileges.')
        return redirect('dashboard')

    try:
        school = request.user.school if request.user.role != 'super_admin' else None

        # Get report cards based on user role
        report_cards = ReportCard.objects.select_related('student', 'grading_period', 'template', 'school')
        
        if request.user.role == 'super_admin':
            # Super admin can see all report cards
            pass
        elif request.user.role == 'admin':
            # Admin can only see report cards from their school
            report_cards = report_cards.filter(school=school)
        elif request.user.role == 'teacher':
            # Teachers can only see report cards for students in their classes
            student_ids = StudentEnrollment.objects.filter(
                class_section__teacher=request.user
            ).values_list('student_id', flat=True)
            report_cards = report_cards.filter(student_id__in=student_ids, school=school)
        elif request.user.role == 'student':
            # Students can only see their own report cards
            report_cards = report_cards.filter(student=request.user)

        # Filter by student if specified
        student_id = request.GET.get('student')
        if student_id:
            try:
                student = User.objects.get(id=student_id, role='student')
                # Verify user has permission to view this student's report cards
                if request.user.role == 'admin' and student.school != request.user.school:
                    messages.error(request, 'Access denied. Cannot view report cards for students from other schools.')
                    return redirect('report_card_list')
                elif request.user.role == 'teacher':
                    if not StudentEnrollment.objects.filter(
                        student=student,
                        class_section__teacher=request.user
                    ).exists() or student.school != request.user.school:
                        messages.error(request, 'Access denied. Cannot view report cards for students you do not teach.')
                        return redirect('report_card_list')
                
                report_cards = report_cards.filter(student_id=student_id)
            except User.DoesNotExist:
                messages.error(request, 'Student not found.')
                return redirect('report_card_list')

        # Filter by grading period if specified
        grading_period_id = request.GET.get('grading_period')
        if grading_period_id:
            try:
                grading_period = GradingPeriod.objects.get(id=grading_period_id)
                # Verify user has permission to view this grading period's report cards
                if request.user.role == 'admin' and grading_period.school != request.user.school:
                    messages.error(request, 'Access denied. Cannot view report cards for grading periods from other schools.')
                    return redirect('report_card_list')
                elif request.user.role == 'teacher':
                    if grading_period.school != request.user.school:
                        messages.error(request, 'Access denied. Cannot view report cards for grading periods from other schools.')
                        return redirect('report_card_list')
                
                report_cards = report_cards.filter(grading_period_id=grading_period_id)
            except GradingPeriod.DoesNotExist:
                messages.error(request, 'Grading period not found.')
                return redirect('report_card_list')

        # Filter by status if specified
        status_filter = request.GET.get('status')
        if status_filter:
            if status_filter in ['draft', 'published', 'archived']:
                report_cards = report_cards.filter(status=status_filter)
            else:
                messages.error(request, 'Invalid status filter.')
                return redirect('report_card_list')

        # Get available students for filtering
        students = User.objects.filter(role='student')
        if request.user.role == 'admin':
            students = students.filter(school=school)
        elif request.user.role == 'teacher':
            student_ids = StudentEnrollment.objects.filter(
                class_section__teacher=request.user
            ).values_list('student_id', flat=True).distinct()
            students = students.filter(id__in=student_ids)
        elif request.user.role == 'student':
            students = students.filter(id=request.user.id)

        # Get available grading periods
        grading_periods = GradingPeriod.objects.filter(school=school) if school else GradingPeriod.objects.all()
        if request.user.role == 'teacher':
            # Get grading periods that have grades from this teacher's subjects
            teacher_subjects = Subject.objects.filter(class_sections__teacher=request.user)
            grading_periods = GradingPeriod.objects.filter(
                grades__subject__in=teacher_subjects
            ).distinct()
        elif request.user.role == 'student':
            # Get grading periods that have grades for this student
            grading_periods = GradingPeriod.objects.filter(
                grades__student=request.user
            ).distinct()

        # Get available classes for filtering
        class_sections = ClassSection.objects.filter(school=school) if school else ClassSection.objects.all()
        if request.user.role == 'teacher':
            class_sections = class_sections.filter(teacher=request.user)
        elif request.user.role == 'student':
            class_sections = ClassSection.objects.filter(
                enrollments__student=request.user
            ).distinct()

        # Pagination
        from django.core.paginator import Paginator
        paginator = Paginator(report_cards, 20)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context = {
            'report_cards': page_obj,
            'students': students.order_by('last_name', 'first_name'),
            'grading_periods': grading_periods,
            'class_sections': class_sections,
            'selected_student_id': student_id,
            'selected_grading_period_id': grading_period_id,
            'selected_status': status_filter,
            'title': 'Report Cards'
        }

        return render(request, 'report_cards/report_card_list.html', context)
    except Exception as e:
        messages.error(request, f'An error occurred: {str(e)}')
        return redirect('dashboard')


@login_required
def report_card_generate(request):
    """Generate report cards for students"""
    if request.user.role not in ['super_admin', 'admin', 'teacher']:
        messages.error(request, 'Access denied. Insufficient privileges.')
        return redirect('dashboard')

    school = request.user.school if request.user.role != 'super_admin' else None

    if request.method == 'POST':
        student_ids = request.POST.getlist('student_ids')
        grading_period_id = request.POST.get('grading_period')
        template_id = request.POST.get('template')
        
        if not student_ids or not grading_period_id or not template_id:
            messages.error(request, 'Please select students, grading period, and template.')
            return redirect('report_card_generate')

        try:
            grading_period = GradingPeriod.objects.get(id=grading_period_id)
            template = ReportTemplate.objects.get(id=template_id, school=school)
            
            # Validate template is active
            if not template.is_active:
                messages.error(request, 'Selected template is not active.')
                return redirect('report_card_generate')

            success_count = 0
            error_count = 0

            for student_id in student_ids:
                try:
                    student = User.objects.get(id=student_id, role='student')
                    
                    # Check if student belongs to school
                    if student.school != school:
                        error_count += 1
                        continue

                    # Check if student has grades for this grading period
                    grades = Grade.objects.filter(
                        student=student,
                        grading_period=grading_period
                    )
                    
                    if not grades.exists():
                        messages.warning(request, f'Student {student.get_full_name()} has no grades for this grading period.')
                        continue

                    # Create or update report card
                    report_card, created = ReportCard.objects.get_or_create(
                        student=student,
                        grading_period=grading_period,
                        template=template,
                        defaults={
                            'school': school,
                            'created_by': request.user
                        }
                    )

                    # Generate report card data
                    report_card.generate_data()
                    
                    # Calculate average grade
                    report_card.calculate_average_grade()
                    
                    # Calculate class rank
                    report_card.get_class_rank()
                    
                    success_count += 1

                except User.DoesNotExist:
                    error_count += 1
                    continue
                except Exception as e:
                    error_count += 1
                    continue

            if success_count > 0:
                messages.success(request, f'Successfully generated {success_count} report cards.')
            if error_count > 0:
                messages.warning(request, f'Failed to generate {error_count} report cards.')

            return redirect('report_card_list')

        except (GradingPeriod.DoesNotExist, ReportTemplate.DoesNotExist) as e:
            messages.error(request, 'Invalid grading period or template selected.')
            return redirect('report_card_generate')

    # Get available students
    students = User.objects.filter(role='student')
    if request.user.role == 'admin':
        students = students.filter(school=school)
    elif request.user.role == 'teacher':
        student_ids = StudentEnrollment.objects.filter(
            class_section__teacher=request.user
        ).values_list('student_id', flat=True).distinct()
        students = students.filter(id__in=student_ids)

    # Get available grading periods
    grading_periods = GradingPeriod.objects.filter(school=school) if school else GradingPeriod.objects.all()
    if request.user.role == 'teacher':
        # Get grading periods that have grades from this teacher's subjects
        teacher_subjects = Subject.objects.filter(class_sections__teacher=request.user)
        grading_periods = GradingPeriod.objects.filter(
            grades__subject__in=teacher_subjects
        ).distinct()

    # Get available templates
    templates = ReportTemplate.objects.filter(school=school, is_active=True) if school else ReportTemplate.objects.filter(is_active=True)

    context = {
        'students': students.order_by('last_name', 'first_name'),
        'grading_periods': grading_periods,
        'templates': templates,
        'title': 'Generate Report Cards'
    }

    return render(request, 'report_cards/report_card_generate.html', context)


@login_required
def publish_report_card(request, report_card_id):
    """Publish a report card"""
    if request.user.role not in ['super_admin', 'admin', 'teacher']:
        messages.error(request, 'Access denied. Insufficient privileges.')
        return redirect('dashboard')

    report_card = get_object_or_404(ReportCard, id=report_card_id)
    
    # Check permissions
    if request.user.role == 'admin' and report_card.school != request.user.school:
        messages.error(request, 'Access denied. Cannot publish report cards from other schools.')
        return redirect('report_card_list')
    elif request.user.role == 'teacher':
        # Check if teacher can publish this student's report card
        if not StudentEnrollment.objects.filter(
            student=report_card.student,
            class_section__teacher=request.user
        ).exists():
            messages.error(request, 'Access denied. Cannot publish report cards for students you do not teach.')
            return redirect('report_card_list')
        elif report_card.school != request.user.school:
            messages.error(request, 'Access denied. Cannot publish report cards from other schools.')
            return redirect('report_card_list')

    # Publish the report card
    report_card.publish(published_by=request.user)
    messages.success(request, f'Report card for {report_card.student.get_full_name()} published successfully.')
    
    return redirect('report_card_list')


@login_required
def unpublish_report_card(request, report_card_id):
    """Unpublish a report card"""
    if request.user.role not in ['super_admin', 'admin', 'teacher']:
        messages.error(request, 'Access denied. Insufficient privileges.')
        return redirect('dashboard')

    report_card = get_object_or_404(ReportCard, id=report_card_id)
    
    # Check permissions (same as publish)
    if request.user.role == 'admin' and report_card.school != request.user.school:
        messages.error(request, 'Access denied. Cannot unpublish report cards from other schools.')
        return redirect('report_card_list')
    elif request.user.role == 'teacher':
        if not StudentEnrollment.objects.filter(
            student=report_card.student,
            class_section__teacher=request.user
        ).exists():
            messages.error(request, 'Access denied. Cannot unpublish report cards for students you do not teach.')
            return redirect('report_card_list')
        elif report_card.school != request.user.school:
            messages.error(request, 'Access denied. Cannot unpublish report cards from other schools.')
            return redirect('report_card_list')

    # Unpublish the report card
    report_card.unpublish()
    messages.success(request, f'Report card for {report_card.student.get_full_name()} unpublished successfully.')
    
    return redirect('report_card_list')


@login_required
def delete_report_card(request, report_card_id):
    """Delete a report card"""
    if request.user.role not in ['super_admin', 'admin', 'teacher']:
        messages.error(request, 'Access denied. Insufficient privileges.')
        return redirect('dashboard')

    report_card = get_object_or_404(ReportCard, id=report_card_id)
    
    # Check permissions (same as publish)
    if request.user.role == 'admin' and report_card.school != request.user.school:
        messages.error(request, 'Access denied. Cannot delete report cards from other schools.')
        return redirect('report_card_list')
    elif request.user.role == 'teacher':
        if not StudentEnrollment.objects.filter(
            student=report_card.student,
            class_section__teacher=request.user
        ).exists():
            messages.error(request, 'Access denied. Cannot delete report cards for students you do not teach.')
            return redirect('report_card_list')
        elif report_card.school != request.user.school:
            messages.error(request, 'Access denied. Cannot delete report cards from other schools.')
            return redirect('report_card_list')

    student_name = report_card.student.get_full_name()
    grading_period_name = report_card.grading_period.name
    
    if request.method == 'POST':
        report_card.delete()
        messages.success(request, f'Report card for {student_name} ({grading_period_name}) deleted successfully.')
        return redirect('report_card_list')

    context = {
        'report_card': report_card,
        'student_name': student_name,
        'grading_period_name': grading_period_name,
        'title': 'Delete Report Card'
    }

    return render(request, 'report_cards/report_card_confirm_delete.html', context)


@login_required
def export_report_cards_pdf(request):
    """Export report cards to PDF"""
    if request.user.role not in ['super_admin', 'admin', 'teacher']:
        messages.error(request, 'Access denied. Insufficient privileges.')
        return redirect('dashboard')

    school = request.user.school if request.user.role != 'super_admin' else None

    # Get selected report cards
    report_card_ids = request.GET.getlist('report_card_ids')
    if not report_card_ids:
        messages.error(request, 'Please select report cards to export.')
        return redirect('report_card_list')

    # Get report cards with proper permissions
    report_cards = ReportCard.objects.filter(id__in=report_card_ids)
    
    if request.user.role == 'admin':
        report_cards = report_cards.filter(school=school)
    elif request.user.role == 'teacher':
        student_ids = StudentEnrollment.objects.filter(
            class_section__teacher=request.user
        ).values_list('student_id', flat=True)
        report_cards = report_cards.filter(student_id__in=student_ids, school=school)
    else:
        report_cards = report_cards.filter(student=request.user)

    if not report_cards.exists():
        messages.error(request, 'No report cards found or access denied.')
        return redirect('report_card_list')

    # Generate PDF
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from django.http import HttpResponse
    from django.conf import settings
    import io
    import os

    # Create PDF buffer
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    story = []

    # Get school profile for branding
    try:
        school_profile = SchoolProfile.objects.get(school=school)
        report_header = school_profile.report_header or f"{school.name}"
        report_footer = school_profile.report_footer or "School Administration"
        report_signature = school_profile.report_signature or "Authorized by Principal"
    except SchoolProfile.DoesNotExist:
        report_header = school.name if school else "ReportCardApp"
        report_footer = "School Administration"
        report_signature = "Authorized by Principal"

    first_report = True
    for report_card in report_cards:
        if not first_report:
            story.append(PageBreak())

        # Get student data
        enrollment = StudentEnrollment.objects.filter(
            student=report_card.student,
            class_section__enrollments__grading_period=report_card.grading_period
        ).first()
        
        grades = report_card.get_grades_data()

        # Header with school branding
        story.append(Paragraph(f"<b>{report_header}</b>", styles['Title']))
        story.append(Paragraph("<b>Report Card</b>", styles['Heading1']))
        story.append(Spacer(1, 12))

        # Student info
        story.append(Paragraph(f"<b>Student Name:</b> {report_card.student.get_full_name()}", styles['Normal']))
        story.append(Paragraph(f"<b>Student ID:</b> {report_card.student.username}", styles['Normal']))
        if enrollment:
            story.append(Paragraph(f"<b>Class:</b> {enrollment.class_section.name}", styles['Normal']))
        story.append(Paragraph(f"<b>Grading Period:</b> {report_card.grading_period.name}", styles['Normal']))
        story.append(Paragraph(f"<b>Academic Year:</b> {report_card.academic_year}", styles['Normal']))
        story.append(Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%B %d, %Y')}", styles['Normal']))
        story.append(Spacer(1, 20))

        # Grades table
        if grades:
            data = [['Subject', 'Score', 'Grade', 'Comments']]
            for grade in grades:
                data.append([
                    grade.subject.name,
                    str(grade.score) if grade.score else '-',
                    grade.letter_grade or '-',
                    grade.comments or '-'
                ])

            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
        else:
            story.append(Paragraph("No grades available.", styles['Normal']))

        # Attendance summary
        attendance_data = report_card.get_attendance_data()
        story.append(Spacer(1, 10))
        story.append(Paragraph("<b>Attendance Summary:</b>", styles['Heading3']))
        story.append(Paragraph(f"Total Days: {attendance_data['total_days']}", styles['Normal']))
        story.append(Paragraph(f"Present: {attendance_data['present_days']}", styles['Normal']))
        story.append(Paragraph(f"Absent: {attendance_data['absent_days']}", styles['Normal']))
        story.append(Paragraph(f"Late: {attendance_data['late_days']}", styles['Normal']))
        story.append(Paragraph(f"Excused: {attendance_data['excused_days']}", styles['Normal']))
        story.append(Paragraph(f"Attendance Rate: {attendance_data['attendance_percentage']}%", styles['Normal']))

        # Average grade and rank
        if report_card.average_grade:
            story.append(Spacer(1, 10))
            story.append(Paragraph(f"<b>Average Grade:</b> {report_card.average_grade}%", styles['Normal']))
        if report_card.class_rank:
            story.append(Paragraph(f"<b>Class Rank:</b> {report_card.class_rank}", styles['Normal']))

        story.append(Spacer(1, 30))
        story.append(Paragraph(f"<i>{report_footer}</i>", styles['Italic']))
        story.append(Spacer(1, 10))
        story.append(Paragraph(f"<i>{report_signature}</i>", styles['Italic']))

        first_report = False

    doc.build(story)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="report_cards.pdf"'

    return response


@login_required
def export_report_cards_excel(request):
    """Export report cards to Excel"""
    if request.user.role not in ['super_admin', 'admin', 'teacher']:
        messages.error(request, 'Access denied. Insufficient privileges.')
        return redirect('dashboard')

    school = request.user.school if request.user.role != 'super_admin' else None

    # Get selected report cards
    report_card_ids = request.GET.getlist('report_card_ids')
    if not report_card_ids:
        messages.error(request, 'Please select report cards to export.')
        return redirect('report_card_list')

    # Get report cards with proper permissions
    report_cards = ReportCard.objects.filter(id__in=report_card_ids)
    
    if request.user.role == 'admin':
        report_cards = report_cards.filter(school=school)
    elif request.user.role == 'teacher':
        student_ids = StudentEnrollment.objects.filter(
            class_section__teacher=request.user
        ).values_list('student_id', flat=True)
        report_cards = report_cards.filter(student_id__in=student_ids, school=school)
    else:
        report_cards = report_cards.filter(student=request.user)

    if not report_cards.exists():
        messages.error(request, 'No report cards found or access denied.')
        return redirect('report_card_list')

    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Report Cards'

    # Headers
    headers = ['Student Name', 'Student ID', 'Grading Period', 'Academic Year', 'Average Grade', 'Class Rank', 'Status', 'Generated At']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Data
    for row_num, report_card in enumerate(report_cards, 2):
        ws.cell(row=row_num, column=1, value=report_card.student.get_full_name())
        ws.cell(row=row_num, column=2, value=report_card.student.username)
        ws.cell(row=row_num, column=3, value=report_card.grading_period.name)
        ws.cell(row=row_num, column=4, value=report_card.academic_year or '')
        ws.cell(row=row_num, column=5, value=report_card.average_grade or '')
        ws.cell(row=row_num, column=6, value=report_card.class_rank or '')
        ws.cell(row=row_num, column=7, value=report_card.status.title())
        ws.cell(row=row_num, column=8, value=report_card.created_at.strftime('%Y-%m-%d %H:%M:%S'))

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=report_cards.xlsx'
    wb.save(response)
    return response


# Analytics Dashboard
@login_required
def analytics_dashboard(request):
    """Analytics dashboard showing grade distributions, attendance trends, and performance metrics"""
    if request.user.role not in ['super_admin', 'admin', 'teacher']:
        messages.error(request, 'Access denied. Insufficient privileges.')
        return redirect('dashboard')

    school = request.user.school if request.user.role != 'super_admin' else None
    
    # Check if school has analytics enabled
    if school:
        try:
            school_profile = SchoolProfile.objects.get(school=school)
            if not school_profile.enable_analytics:
                messages.warning(request, 'Analytics is not enabled for this school.')
                return redirect('dashboard')
        except SchoolProfile.DoesNotExist:
            pass

    from django.db.models import Q, Count, Avg, Min, Max
    
    # Get grade statistics
    grades_qs = Grade.objects.select_related('student', 'subject', 'grading_period')
    if request.user.role == 'admin':
        grades_qs = grades_qs.filter(school=school)
    elif request.user.role == 'teacher':
        grades_qs = grades_qs.filter(school=request.user.school, subject__class_sections__teacher=request.user).distinct()

    # Grade distribution by letter grade
    grade_distribution = {}
    for grade in grades_qs:
        letter = grade.letter_grade or 'N/A'
        grade_distribution[letter] = grade_distribution.get(letter, 0) + 1

    # Score statistics
    score_stats = grades_qs.aggregate(
        avg_score=Avg('score'),
        min_score=Min('score'),
        max_score=Max('score'),
        total_grades=Count('id')
    )

    # Top performing students (by average score)
    from django.db.models import Avg as AvgFunc
    top_students = grades_qs.values('student__id', 'student__first_name', 'student__last_name', 'student__username').annotate(
        avg_score=AvgFunc('score')
    ).filter(avg_score__isnull=False).order_by('-avg_score')[:10]

    # Performance by subject
    subject_performance = grades_qs.values('subject__id', 'subject__name').annotate(
        avg_score=AvgFunc('score'),
        count=Count('id')
    ).filter(avg_score__isnull=False).order_by('-avg_score')

    # Attendance statistics
    attendance_qs = Attendance.objects.select_related('student', 'class_section')
    if request.user.role == 'admin':
        attendance_qs = attendance_qs.filter(school=school)
    elif request.user.role == 'teacher':
        attendance_qs = attendance_qs.filter(school=request.user.school, class_section__teacher=request.user)

    attendance_stats = attendance_qs.aggregate(
        total_records=Count('id'),
        present_count=Count('id', filter=Q(status='present')),
        absent_count=Count('id', filter=Q(status='absent')),
        late_count=Count('id', filter=Q(status='late')),
        excused_count=Count('id', filter=Q(status='excused'))
    )

    # Calculate attendance percentage
    if attendance_stats['total_records'] > 0:
        attendance_stats['present_percentage'] = round((attendance_stats['present_count'] / attendance_stats['total_records']) * 100, 2)
        attendance_stats['absent_percentage'] = round((attendance_stats['absent_count'] / attendance_stats['total_records']) * 100, 2)
    else:
        attendance_stats['present_percentage'] = 0
        attendance_stats['absent_percentage'] = 0

    # Students with low attendance (less than 80%)
    from django.db.models import Case, When, FloatField, F
    from django.db.models.functions import Cast
    low_attendance_students = attendance_qs.values('student__id', 'student__first_name', 'student__last_name').annotate(
        total=Count('id'),
        present=Count('id', filter=Q(status='present'))
    ).annotate(
        attendance_pct=Case(
            When(total=0, then=0),
            default=Cast(F('present') * 100.0 / F('total'), output_field=FloatField())
        )
    ).filter(attendance_pct__lt=80).order_by('attendance_pct')[:10]

    # Grading periods for filtering
    grading_periods = GradingPeriod.objects.filter(school=school) if school else GradingPeriod.objects.all()
    if request.user.role == 'teacher':
        # Get grading periods that have grades from this teacher's subjects
        teacher_subjects = Subject.objects.filter(class_sections__teacher=request.user)
        grading_periods = GradingPeriod.objects.filter(
            grades__subject__in=teacher_subjects
        ).distinct()

    # Filter by grading period if specified
    selected_period_id = request.GET.get('grading_period')
    if selected_period_id:
        grades_qs = grades_qs.filter(grading_period_id=selected_period_id)
        attendance_qs = attendance_qs.filter(date__gte=GradingPeriod.objects.get(id=selected_period_id).start_date,
                                             date__lte=GradingPeriod.objects.get(id=selected_period_id).end_date)

    context = {
        'grade_distribution': grade_distribution,
        'score_stats': score_stats,
        'top_students': top_students,
        'subject_performance': subject_performance,
        'attendance_stats': attendance_stats,
        'low_attendance_students': low_attendance_students,
        'grading_periods': grading_periods,
        'selected_period_id': selected_period_id,
        'school': school,
        'title': 'Analytics Dashboard'
    }

    return render(request, 'analytics/dashboard.html', context)




# Global Search View
@login_required
def search_view(request):
    query = request.GET.get('q', '').strip()
    results = {
        'schools': [],
        'users': [],
        'classes': [],
        'subjects': [],
        'grades': [],
        'attendances': [],
    }

    if query:
        # Search schools
        if request.user.role == 'super_admin':
            results['schools'] = School.objects.filter(name__icontains=query)
        else:
            results['schools'] = School.objects.filter(name__icontains=query, id=request.user.school.id) if request.user.school else []

        # Search users
        if request.user.role == 'super_admin':
            results['users'] = User.objects.filter(
                Q(username__icontains=query) |
                Q(first_name__icontains=query) |
                Q(last_name__icontains=query) |
                Q(email__icontains=query)
            )[:20]
        else:
            results['users'] = User.objects.filter(
                Q(username__icontains=query) |
                Q(first_name__icontains=query) |
                Q(last_name__icontains=query) |
                Q(email__icontains=query),
                school=request.user.school
            )[:20]

        # Search classes
        class_sections = ClassSection.objects.filter(
            Q(name__icontains=query) |
            Q(grade_level__icontains=query)
        )
        if request.user.role == 'admin':
            class_sections = class_sections.filter(school=request.user.school)
        elif request.user.role == 'teacher':
            class_sections = class_sections.filter(teacher=request.user)
        results['classes'] = class_sections[:20]

        # Search subjects
        subjects = Subject.objects.filter(
            Q(name__icontains=query) |
            Q(code__icontains=query) |
            Q(description__icontains=query)
        )
        if request.user.role in ['admin', 'teacher']:
            subjects = subjects.filter(school=request.user.school)
        results['subjects'] = subjects[:20]

        # Search grades (for teachers and admins)
        if request.user.role in ['super_admin', 'admin', 'teacher']:
            grades = Grade.objects.filter(
                Q(letter_grade__icontains=query) |
                Q(comments__icontains=query)
            ).select_related('student', 'subject', 'grading_period')
            if request.user.role == 'admin':
                grades = grades.filter(school=request.user.school)
            elif request.user.role == 'teacher':
                grades = grades.filter(school=request.user.school, subject__class_sections__teacher=request.user).distinct()
            results['grades'] = grades[:20]

        # Search attendance (for teachers and admins)
        if request.user.role in ['super_admin', 'admin', 'teacher']:
            attendances = Attendance.objects.filter(
                Q(notes__icontains=query)
            ).select_related('student', 'class_section')
            if request.user.role == 'admin':
                attendances = attendances.filter(school=request.user.school)
            elif request.user.role == 'teacher':
                attendances = attendances.filter(school=request.user.school, class_section__teacher=request.user)
            results['attendances'] = attendances[:20]

    return render(request, 'schools/search.html', {
        'query': query,
        'results': results,
        'title': 'Search Results'
    })


# Export Views
@login_required
def export_grades_excel(request):
    if not PermissionHelper.user_can_export(request.user):
        return HttpResponse('Unauthorized', status=403)

    school = PermissionHelper.get_user_school(request.user)
    grades = Grade.objects.select_related('student', 'subject', 'grading_period', 'school')
    
    if request.user.role == 'admin':
        grades = grades.filter(school=school)
    elif request.user.role == 'teacher':
        grades = grades.filter(school=request.user.school, subject__class_sections__teacher=request.user).distinct()

    exporter = ExcelExporter(
        'Grades',
        ['Student ID', 'Student Name', 'Subject', 'Grading Period', 'Score', 'Letter Grade', 'Comments', 'School']
    )
    
    for grade in grades:
        exporter.add_row([
            grade.student.username, grade.student.get_full_name(), grade.subject.name,
            grade.grading_period.name, grade.score, grade.letter_grade, grade.comments, grade.school.name
        ])
    
    return exporter.get_response('grades')


@login_required
def export_attendance_excel(request):
    if not PermissionHelper.user_can_export(request.user):
        return HttpResponse('Unauthorized', status=403)

    school = PermissionHelper.get_user_school(request.user)
    attendances = Attendance.objects.select_related('student', 'class_section', 'school')
    
    if request.user.role == 'admin':
        attendances = attendances.filter(school=school)
    elif request.user.role == 'teacher':
        attendances = attendances.filter(school=request.user.school, class_section__teacher=request.user)

    exporter = ExcelExporter(
        'Attendance',
        ['Student ID', 'Student Name', 'Class Section', 'Date', 'Status', 'Notes', 'School']
    )
    
    for attendance in attendances:
        exporter.add_row([
            attendance.student.username, attendance.student.get_full_name(), attendance.class_section.name,
            str(attendance.date), attendance.status, attendance.notes, attendance.school.name
        ])
    
    return exporter.get_response('attendance')


@login_required
def export_users_csv(request):
    if request.user.role not in ['super_admin', 'admin']:
        return HttpResponse('Unauthorized', status=403)

    school = PermissionHelper.get_user_school(request.user)
    users = User.objects.all()
    if school:
        users = users.filter(school=school)

    exporter = CSVExporter('users', ['ID', 'Username', 'First Name', 'Last Name', 'Email', 'Role', 'School'])
    
    for user in users:
        exporter.add_row([
            user.id, user.username, user.first_name, user.last_name, user.email,
            user.role, user.school.name if user.school else ''
        ])
    
    return exporter.get_response()


# School Profile Management Views (White-Label Features)
@login_required
def school_profile_view(request):
    """View and edit school branding and white-label settings"""
    if request.user.role not in ['super_admin', 'admin']:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboard')

    school = request.user.school if request.user.role == 'admin' else None
    
    # Get or create school profile
    school_profile, created = SchoolProfile.objects.get_or_create(
        school=request.user.school,
        defaults={
            'primary_color': '#667eea',
            'secondary_color': '#764ba2',
            'accent_color': '#28a745'
        }
    )

    if request.method == 'POST':
        from .forms import SchoolProfileForm
        form = SchoolProfileForm(request.POST, request.FILES, instance=school_profile)
        if form.is_valid():
            form.save()
            messages.success(request, 'School profile updated successfully.')
            return redirect('school_profile')
    else:
        form = SchoolProfileForm(instance=school_profile)

    return render(request, 'schools/school_profile.html', {
        'form': form,
        'school_profile': school_profile,
        'school': request.user.school,
        'title': 'School Branding & Settings'
    })


# Support Ticket System Views
@login_required
def support_ticket_list(request):
    """List support tickets for the user"""
    tickets = SupportTicket.objects.filter(created_by=request.user).order_by('-created_at')
    
    # Admins and super admins can see all tickets for their school
    if request.user.role in ['admin', 'super_admin']:
        if request.user.role == 'super_admin':
            tickets = SupportTicket.objects.all().order_by('-created_at')
        else:
            tickets = SupportTicket.objects.filter(school=request.user.school).order_by('-created_at')

    return render(request, 'support/ticket_list.html', {
        'tickets': tickets,
        'title': 'Support Tickets'
    })


@login_required
def support_ticket_create(request):
    """Create a new support ticket"""
    if request.method == 'POST':
        from .forms import SupportTicketForm
        form = SupportTicketForm(request.POST, request=request)
        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.created_by = request.user
            ticket.school = request.user.school
            ticket.save()
            messages.success(request, 'Support ticket created successfully. We will get back to you soon.')
            return redirect('support_ticket_list')
    else:
        form = SupportTicketForm(request=request)

    return render(request, 'support/ticket_create.html', {
        'form': form,
        'title': 'Create Support Ticket'
    })


@login_required
def support_ticket_detail(request, pk):
    """View support ticket details"""
    ticket = get_object_or_404(SupportTicket, pk=pk)
    
    # Check permissions
    if request.user.role not in ['super_admin', 'admin']:
        # Regular users can only see their own tickets
        if ticket.created_by != request.user:
            messages.error(request, 'Access denied.')
            return redirect('support_ticket_list')
    else:
        # Admins can see tickets for their school
        if request.user.role == 'admin' and ticket.school != request.user.school:
            messages.error(request, 'Access denied.')
            return redirect('support_ticket_list')

    return render(request, 'support/ticket_detail.html', {
        'ticket': ticket,
        'title': f'Ticket: {ticket.title}'
    })


@login_required
def support_dashboard(request):
    """Admin dashboard for managing support tickets"""
    if request.user.role not in ['super_admin', 'admin']:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboard')

    # Get tickets for the school or all tickets for super admin
    if request.user.role == 'super_admin':
        tickets = SupportTicket.objects.all()
    else:
        tickets = SupportTicket.objects.filter(school=request.user.school)

    # Filter by status if specified
    status_filter = request.GET.get('status')
    if status_filter:
        tickets = tickets.filter(status=status_filter)

    # Filter by priority if specified
    priority_filter = request.GET.get('priority')
    if priority_filter:
        tickets = tickets.filter(priority=priority_filter)

    # Get statistics
    total_tickets = tickets.count()
    open_tickets = tickets.filter(status='open').count()
    in_progress_tickets = tickets.filter(status='in_progress').count()
    resolved_tickets = tickets.filter(status='resolved').count()

    return render(request, 'support/dashboard.html', {
        'tickets': tickets,
        'total_tickets': total_tickets,
        'open_tickets': open_tickets,
        'in_progress_tickets': in_progress_tickets,
        'resolved_tickets': resolved_tickets,
        'status_filter': status_filter,
        'priority_filter': priority_filter,
        'title': 'Support Dashboard'
    })


@login_required
def support_ticket_update(request, pk):
    """Update support ticket (for admins)"""
    if request.user.role not in ['super_admin', 'admin']:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('support_ticket_list')

    ticket = get_object_or_404(SupportTicket, pk=pk)
    
    # Check permissions
    if request.user.role == 'admin' and ticket.school != request.user.school:
        messages.error(request, 'Access denied.')
        return redirect('support_ticket_list')

    if request.method == 'POST':
        from .forms import SupportTicketAdminForm
        form = SupportTicketAdminForm(request.POST, instance=ticket)
        if form.is_valid():
            ticket = form.save()
            messages.success(request, 'Ticket updated successfully.')
            return redirect('support_ticket_detail', pk=ticket.pk)
    else:
        form = SupportTicketAdminForm(instance=ticket)

    return render(request, 'support/ticket_update.html', {
        'form': form,
        'ticket': ticket,
        'title': f'Update Ticket: {ticket.title}'
    })


@login_required
def support_ticket_assign(request, pk):
    """Assign ticket to staff member (for admins)"""
    if request.user.role not in ['super_admin', 'admin']:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('support_ticket_list')

    ticket = get_object_or_404(SupportTicket, pk=pk)
    
    # Check permissions
    if request.user.role == 'admin' and ticket.school != request.user.school:
        messages.error(request, 'Access denied.')
        return redirect('support_ticket_list')

    if request.method == 'POST':
        assigned_to_id = request.POST.get('assigned_to')
        if assigned_to_id:
            try:
                assigned_user = User.objects.get(id=assigned_to_id, role__in=['admin', 'super_admin'])
                ticket.assigned_to = assigned_user
                ticket.save()
                messages.success(request, f'Ticket assigned to {assigned_user.get_full_name()}.')
            except User.DoesNotExist:
                messages.error(request, 'Invalid user selected.')
        return redirect('support_ticket_detail', pk=ticket.pk)

    # Get available staff members
    staff_members = User.objects.filter(role__in=['admin', 'super_admin']).order_by('last_name', 'first_name')
    
    return render(request, 'support/ticket_assign.html', {
        'ticket': ticket,
        'staff_members': staff_members,
        'title': f'Assign Ticket: {ticket.title}'
    })


# User Profile & Settings Views
@login_required
def user_profile(request):
    """Display and manage user profile"""
    user = request.user
    
    if request.method == 'POST':
        form_type = request.POST.get('form_type')
        
        if form_type == 'profile':
            # Update basic profile info
            user.first_name = request.POST.get('first_name', user.first_name)
            user.last_name = request.POST.get('last_name', user.last_name)
            user.email = request.POST.get('email', user.email)
            
            # Handle profile picture upload
            if 'profile_picture' in request.FILES:
                user.profile_picture = request.FILES['profile_picture']
            
            user.save()
            messages.success(request, 'Profile updated successfully.')
            return redirect('user_profile')
    
    context = {
        'user': user,
        'title': 'My Profile'
    }
    return render(request, 'users/user_profile.html', context)


@login_required
def user_settings(request):
    """User settings and preferences"""
    user = request.user
    
    if request.method == 'POST':
        setting_type = request.POST.get('setting_type')
        
        if setting_type == 'change_password':
            # Handle password change
            current_password = request.POST.get('current_password')
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('confirm_password')
            
            if not user.check_password(current_password):
                messages.error(request, 'Current password is incorrect.')
            elif new_password != confirm_password:
                messages.error(request, 'New passwords do not match.')
            elif len(new_password) < 8:
                messages.error(request, 'Password must be at least 8 characters long.')
            else:
                user.set_password(new_password)
                user.save()
                messages.success(request, 'Password changed successfully.')
                return redirect('auth:login')  # Redirect to login after password change
    
    context = {
        'user': user,
        'title': 'Settings & Security'
    }
    return render(request, 'users/user_settings.html', context)


@login_required
def help_center(request):
    """Help center and support documentation"""
    
    context = {
        'title': 'Help & Support'
    }
    return render(request, 'support/help_center.html', context)


# Student Portal Views
@login_required
def student_grades(request):
    """Display student's own grades"""
    if request.user.role != 'student':
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    student = request.user
    grades = Grade.objects.filter(student=student).select_related(
        'subject', 'grading_period'
    ).order_by('-grading_period__end_date')
    
    # Add filtering by grading period if requested
    grading_period_id = request.GET.get('grading_period')
    if grading_period_id:
        grades = grades.filter(grading_period_id=grading_period_id)
    
    grading_periods = GradingPeriod.objects.filter(
        class_section__in=StudentEnrollment.objects.filter(student=student).values_list('class_section_id')
    ).distinct().order_by('-end_date')
    
    context = {
        'grades': grades,
        'grading_periods': grading_periods,
        'student': student,
        'title': 'My Grades'
    }
    return render(request, 'students/student_grades.html', context)


@login_required
def student_attendance(request):
    """Display student's own attendance records"""
    if request.user.role != 'student':
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    student = request.user
    attendance_records = Attendance.objects.filter(student=student).select_related(
        'class_section'
    ).order_by('-date')
    
    # Calculate attendance statistics
    total_sessions = attendance_records.count()
    present_count = attendance_records.filter(status='present').count()
    absent_count = attendance_records.filter(status='absent').count()
    late_count = attendance_records.filter(status='late').count()
    
    attendance_rate = (present_count / total_sessions * 100) if total_sessions > 0 else 0
    
    context = {
        'attendance_records': attendance_records,
        'total_sessions': total_sessions,
        'present_count': present_count,
        'absent_count': absent_count,
        'late_count': late_count,
        'attendance_rate': attendance_rate,
        'student': student,
        'title': 'My Attendance'
    }
    return render(request, 'students/student_attendance.html', context)


@login_required
def student_report_cards(request):
    """Display student's own report cards"""
    if request.user.role != 'student':
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    
    student = request.user
    report_cards = ReportCard.objects.filter(student=student).select_related(
        'grading_period'
    ).order_by('-created_at')
    
    # Filter by grading period if requested
    grading_period_id = request.GET.get('grading_period')
    if grading_period_id:
        report_cards = report_cards.filter(grading_period_id=grading_period_id)
    
    grading_periods = GradingPeriod.objects.filter(
        reportcard__student=student
    ).distinct().order_by('-end_date')
    
    context = {
        'report_cards': report_cards,
        'grading_periods': grading_periods,
        'student': student,
        'title': 'My Report Cards'
    }
    return render(request, 'students/student_report_cards.html', context)
