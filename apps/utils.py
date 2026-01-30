"""
Common utilities and helper functions to eliminate code duplication.
Includes export formatters, permission checks, and data processing functions.
"""
import csv
import io
import json
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

from django.http import HttpResponse
from django.db.models import Count, Q, Avg, Min, Max, Case, When, FloatField, F
from django.db.models.functions import Cast

from apps.models import StudentEnrollment


class ExcelExporter:
    """Unified Excel export formatter"""
    
    def __init__(self, title="Export", headers=None):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = title
        self.headers = headers or []
        self.current_row = 1
        self._write_headers()
    
    def _write_headers(self):
        """Write header row with formatting"""
        for col_num, header in enumerate(self.headers, 1):
            cell = self.ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    def add_row(self, data):
        """Add a data row"""
        self.current_row += 1
        for col_num, value in enumerate(data, 1):
            cell = self.ws.cell(row=self.current_row, column=col_num, value=value)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    def auto_adjust_columns(self):
        """Auto-adjust column widths"""
        for column in self.ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            self.ws.column_dimensions[column_letter].width = adjusted_width
    
    def get_response(self, filename):
        """Get HttpResponse for download"""
        self.auto_adjust_columns()
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        self.wb.save(response)
        return response


class PDFExporter:
    """Unified PDF export formatter using ReportLab"""
    
    def __init__(self, title="Export"):
        self.buffer = io.BytesIO()
        self.doc = SimpleDocTemplate(
            self.buffer,
            pagesize=letter,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        self.styles = getSampleStyleSheet()
        self.story = []
        self.title = title
    
    def add_title(self, text):
        """Add title"""
        self.story.append(Paragraph(f"<b>{text}</b>", self.styles['Title']))
        self.story.append(Spacer(1, 12))
    
    def add_heading(self, text):
        """Add section heading"""
        self.story.append(Paragraph(f"<b>{text}</b>", self.styles['Heading2']))
        self.story.append(Spacer(1, 10))
    
    def add_paragraph(self, text):
        """Add paragraph"""
        self.story.append(Paragraph(text, self.styles['Normal']))
    
    def add_table(self, data, style=None):
        """Add table"""
        table = Table(data)
        if not style:
            style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ])
        table.setStyle(style)
        self.story.append(table)
    
    def add_spacer(self, height=12):
        """Add spacing"""
        self.story.append(Spacer(1, height))
    
    def add_page_break(self):
        """Add page break"""
        self.story.append(PageBreak())
    
    def get_response(self, filename):
        """Build and return response"""
        self.doc.build(self.story)
        self.buffer.seek(0)
        response = HttpResponse(self.buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        return response


class CSVExporter:
    """Unified CSV export formatter"""
    
    def __init__(self, filename, headers):
        self.response = HttpResponse(content_type='text/csv')
        self.response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        self.writer = csv.writer(self.response)
        self.writer.writerow(headers)
    
    def add_row(self, data):
        """Add a row"""
        self.writer.writerow(data)
    
    def get_response(self):
        """Get response"""
        return self.response


class PermissionHelper:
    """Helper functions for permission checking"""
    
    @staticmethod
    def user_can_export(user):
        """Check if user can export data"""
        return user.role in ['super_admin', 'admin', 'teacher']
    
    @staticmethod
    def get_user_school(user):
        """Get school from user (super admin gets None)"""
        return user.school if user.role != 'super_admin' else None
    
    @staticmethod
    def filter_by_school(queryset, user):
        """Apply school filtering based on user role"""
        if user.role == 'super_admin':
            return queryset
        elif user.role in ['admin', 'teacher']:
            school = PermissionHelper.get_user_school(user)
            if school:
                return queryset.filter(school=school)
        return queryset.none()
    
    @staticmethod
    def filter_teacher_students(user):
        """Get student IDs for a teacher's classes"""
        if user.role != 'teacher':
            return StudentEnrollment.objects.none()
        
        return StudentEnrollment.objects.filter(
            class_section__teacher=user
        ).values_list('student_id', flat=True).distinct()


class AnalyticsHelper:
    """Helper functions for analytics calculations"""
    
    @staticmethod
    def get_attendance_stats(attendance_qs):
        """Calculate attendance statistics from queryset"""
        stats = attendance_qs.aggregate(
            total_records=Count('id'),
            present_count=Count('id', filter=Q(status='present')),
            absent_count=Count('id', filter=Q(status='absent')),
            late_count=Count('id', filter=Q(status='late')),
            excused_count=Count('id', filter=Q(status='excused'))
        )
        
        if stats['total_records'] > 0:
            stats['present_percentage'] = round((stats['present_count'] / stats['total_records']) * 100, 2)
            stats['absent_percentage'] = round((stats['absent_count'] / stats['total_records']) * 100, 2)
        else:
            stats['present_percentage'] = 0
            stats['absent_percentage'] = 0
        
        return stats
    
    @staticmethod
    def get_grade_distribution(grades_qs):
        """Calculate grade distribution"""
        distribution = {}
        for grade in grades_qs:
            letter = grade.letter_grade or 'N/A'
            distribution[letter] = distribution.get(letter, 0) + 1
        return distribution
    
    @staticmethod
    def get_score_statistics(grades_qs):
        """Get score statistics"""
        return grades_qs.aggregate(
            avg_score=Avg('score'),
            min_score=Min('score'),
            max_score=Max('score'),
            total_grades=Count('id')
        )
    
    @staticmethod
    def get_low_attendance_students(attendance_qs, threshold=80):
        """Get students with attendance below threshold"""
        return attendance_qs.values(
            'student__id', 'student__first_name', 'student__last_name'
        ).annotate(
            total=Count('id'),
            present=Count('id', filter=Q(status='present'))
        ).annotate(
            attendance_pct=Case(
                When(total=0, then=0),
                default=Cast(F('present') * 100.0 / F('total'), output_field=FloatField())
            )
        ).filter(attendance_pct__lt=threshold).order_by('attendance_pct')


class ValidationHelper:
    """Helper functions for common validations"""
    
    @staticmethod
    def validate_school_access(user, obj):
        """Validate that user has access to object based on school"""
        if user.role == 'super_admin':
            return True
        
        if user.role == 'admin':
            obj_school = getattr(obj, 'school', None)
            return obj_school == user.school
        
        return False
    
    @staticmethod
    def validate_student_owner(user, student_obj):
        """Validate that user is the student or can view the student"""
        if user.role == 'super_admin':
            return True
        elif user.role == 'admin':
            return student_obj.school == user.school
        elif user.role == 'student':
            return student_obj == user
        elif user.role == 'teacher':
            return StudentEnrollment.objects.filter(
                student=student_obj,
                class_section__teacher=user
            ).exists()
        
        return False
